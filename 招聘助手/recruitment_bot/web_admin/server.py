#!/usr/bin/env python3
"""
招聘助手 Web 管理后台与 API 服务

功能：
- 保存浏览器插件同步的候选人、推荐摘要、Markdown 报告
- 提供 Web GUI 查看历史数据、配置钉钉、问答 Agent
- 接收钉钉回调并根据历史数据回答
- 定时任务由外部 cron 调用 /api/summary/push?scope=yesterday
"""

from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import ipaddress
import io
import json
import os
import re
import socket
import sqlite3
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
import threading
import http.client
from datetime import datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parent
STATIC_DIR = ROOT / "static"
TEMPLATE_DIR = ROOT / "templates"
EXPORT_DIR = ROOT / "exports"
EXPORT_MANIFEST = EXPORT_DIR / "manifest.json"
EXTENSION_DIR = Path(os.environ.get("RECRUITMENT_EXTENSION_DIR", "")).expanduser() if os.environ.get("RECRUITMENT_EXTENSION_DIR") else (
    ROOT / "browser-extension" if (ROOT / "browser-extension").exists() else PROJECT_ROOT / "browser-extension"
)
RECOMMENDATION_TEMPLATE = TEMPLATE_DIR / "定时推送候选人推荐表模板.xlsx"
DB_PATH = Path(os.environ.get("RECRUITMENT_DB", ROOT / "recruitment_history.db"))
HOST = os.environ.get("RECRUITMENT_HOST", "0.0.0.0")
PORT = int(os.environ.get("RECRUITMENT_PORT", "8787"))
IP_ALLOWLIST_RAW = os.environ.get("RECRUITMENT_IP_ALLOWLIST", "127.0.0.1/32,::1/128")
TRUST_PROXY_HEADERS = os.environ.get("RECRUITMENT_TRUST_PROXY_HEADERS", "0").lower() in {"1", "true", "yes", "on"}
MAX_BODY_BYTES = int(os.environ.get("RECRUITMENT_MAX_BODY_BYTES", str(6 * 1024 * 1024)))
RATE_LIMIT_PER_MINUTE = int(os.environ.get("RECRUITMENT_RATE_LIMIT_PER_MINUTE", "120"))
RATE_LIMIT_WINDOW_SECONDS = 60
DEFAULT_PUBLIC_BASE_URL = os.environ.get(
    "RECRUITMENT_PUBLIC_BASE_URL",
    "https://unconfuted-superbusily-ryan.ngrok-free.dev",
).rstrip("/")
EXPORT_FILENAME_PATTERN = re.compile(r"^候选人推荐表_\d{8}_\d{6}\.xlsx$")
SECURITY_ALLOWLIST_PATH = Path(os.environ.get("RECRUITMENT_SECURITY_ALLOWLIST", ROOT / "security_allowlist.json"))
ADMIN_TOKEN = os.environ.get("RECRUITMENT_ADMIN_TOKEN", "").strip()
RATE_BUCKETS: dict[tuple[str, str], list[float]] = {}
RATE_LOCK = threading.Lock()


def _split_env_list(name: str) -> list[str]:
    return [item.strip() for item in os.environ.get(name, "").split(",") if item.strip()]


def _load_security_allowlist() -> dict[str, Any]:
    if not SECURITY_ALLOWLIST_PATH.exists():
        return {}
    try:
        payload = json.loads(SECURITY_ALLOWLIST_PATH.read_text("utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception as exc:
        print(f"[security] allowlist load failed: {exc}")
        return {}


SECURITY_ALLOWLIST = _load_security_allowlist()
MAX_JSON_BODY_BYTES = int(os.environ.get(
    "RECRUITMENT_MAX_JSON_BODY_BYTES",
    os.environ.get("RECRUITMENT_MAX_BODY_BYTES", str(SECURITY_ALLOWLIST.get("maxJsonBodyBytes") or 1024 * 1024)),
))
RATE_LIMIT_PER_MINUTE = int(os.environ.get(
    "RECRUITMENT_RATE_LIMIT_PER_MINUTE",
    str(SECURITY_ALLOWLIST.get("rateLimitPerMinute") or 240),
))
ALLOWED_HOSTS = sorted({
    str(item).strip().lower()
    for item in [*SECURITY_ALLOWLIST.get("allowedHosts", []), *_split_env_list("RECRUITMENT_ALLOWED_HOSTS")]
    if str(item).strip()
})
ALLOWED_ORIGINS = sorted({
    str(item).strip().rstrip("/")
    for item in [*SECURITY_ALLOWLIST.get("allowedOrigins", []), *_split_env_list("RECRUITMENT_ALLOWED_ORIGINS")]
    if str(item).strip()
})
ALLOWED_CLIENT_IPS = [
    str(item).strip()
    for item in [
        *SECURITY_ALLOWLIST.get("allowedClientIps", []),
        *_split_env_list("RECRUITMENT_ALLOWED_CLIENT_IPS"),
        *_split_env_list("RECRUITMENT_IP_ALLOWLIST"),
    ]
    if str(item).strip()
]

REQUEST_BUCKETS: dict[str, list[float]] = {}


def split_config_list(value: Any) -> list[str]:
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = re.split(r"[\n,，]+", str(value or ""))
    return [str(item).strip() for item in raw_items if str(item).strip()]


def normalize_origin_value(value: Any) -> str:
    return str(value or "").strip().rstrip("/")


def clean_allowed_hosts(values: Any) -> list[str]:
    return sorted({
        normalize_host_header(item)
        for item in split_config_list(values)
        if normalize_host_header(item)
    })


def clean_allowed_origins(values: Any) -> list[str]:
    return sorted({
        normalize_origin_value(item)
        for item in split_config_list(values)
        if normalize_origin_value(item)
    })


def clean_allowed_client_ips(values: Any) -> list[str]:
    cleaned: list[str] = []
    for item in split_config_list(values):
        try:
            cleaned.append(str(ipaddress.ip_network(item, strict=False)))
        except ValueError:
            try:
                cleaned.append(str(ipaddress.ip_address(item)))
            except ValueError:
                raise ValueError(f"无效的客户端 IP/CIDR：{item}") from None
    return sorted(set(cleaned))


def reload_security_runtime(config: dict[str, Any] | None = None) -> None:
    global SECURITY_ALLOWLIST, ALLOWED_HOSTS, ALLOWED_ORIGINS, ALLOWED_CLIENT_IPS, RATE_LIMIT_PER_MINUTE, MAX_JSON_BODY_BYTES
    SECURITY_ALLOWLIST = config if config is not None else _load_security_allowlist()
    ALLOWED_HOSTS = sorted({
        *clean_allowed_hosts(SECURITY_ALLOWLIST.get("allowedHosts", [])),
        *clean_allowed_hosts(_split_env_list("RECRUITMENT_ALLOWED_HOSTS")),
    })
    ALLOWED_ORIGINS = sorted({
        *clean_allowed_origins(SECURITY_ALLOWLIST.get("allowedOrigins", [])),
        *clean_allowed_origins(_split_env_list("RECRUITMENT_ALLOWED_ORIGINS")),
    })
    ALLOWED_CLIENT_IPS = sorted({
        *clean_allowed_client_ips(SECURITY_ALLOWLIST.get("allowedClientIps", [])),
        *clean_allowed_client_ips(_split_env_list("RECRUITMENT_ALLOWED_CLIENT_IPS")),
        *clean_allowed_client_ips(_split_env_list("RECRUITMENT_IP_ALLOWLIST")),
    })
    RATE_LIMIT_PER_MINUTE = int(os.environ.get(
        "RECRUITMENT_RATE_LIMIT_PER_MINUTE",
        str(SECURITY_ALLOWLIST.get("rateLimitPerMinute") or 240),
    ))
    MAX_JSON_BODY_BYTES = int(os.environ.get(
        "RECRUITMENT_MAX_JSON_BODY_BYTES",
        os.environ.get("RECRUITMENT_MAX_BODY_BYTES", str(SECURITY_ALLOWLIST.get("maxJsonBodyBytes") or 1024 * 1024)),
    ))


def parse_ip_allowlist(value: str) -> list[Any]:
    networks: list[Any] = []
    for item in str(value or "").split(","):
        item = item.strip()
        if not item:
            continue
        try:
            networks.append(ipaddress.ip_network(item, strict=False))
        except ValueError:
            print(f"[security] ignored invalid allowlist entry: {item}")
    return networks


IP_ALLOWLIST = parse_ip_allowlist(IP_ALLOWLIST_RAW)


def client_ip_from_handler(handler: SimpleHTTPRequestHandler) -> str:
    remote_ip = handler.client_address[0] if handler.client_address else ""
    if TRUST_PROXY_HEADERS and remote_ip in {"127.0.0.1", "::1"}:
        forwarded = (
            handler.headers.get("CF-Connecting-IP")
            or handler.headers.get("X-Real-IP")
            or handler.headers.get("X-Forwarded-For", "").split(",")[0]
        )
        forwarded = str(forwarded or "").strip()
        if forwarded:
            return forwarded
    return remote_ip


def ip_is_allowed(ip_value: str) -> bool:
    if not IP_ALLOWLIST:
        return True
    try:
        ip_obj = ipaddress.ip_address(ip_value)
    except ValueError:
        return False
    return any(ip_obj in network for network in IP_ALLOWLIST)


def rate_limit_allows(ip_value: str) -> bool:
    if RATE_LIMIT_PER_MINUTE <= 0:
        return True
    now = time.time()
    cutoff = now - RATE_LIMIT_WINDOW_SECONDS
    bucket = REQUEST_BUCKETS.setdefault(ip_value, [])
    bucket[:] = [stamp for stamp in bucket if stamp >= cutoff]
    if len(bucket) >= RATE_LIMIT_PER_MINUTE:
        return False
    bucket.append(now)
    if len(REQUEST_BUCKETS) > 2048:
        for key in list(REQUEST_BUCKETS.keys())[:512]:
            if not REQUEST_BUCKETS[key] or REQUEST_BUCKETS[key][-1] < cutoff:
                REQUEST_BUCKETS.pop(key, None)
    return True


DEFAULT_BEHAVIOR_POLICY: dict[str, Any] = {
    "behaviorPolicyEnabled": True,
    "workTimeEnabled": False,
    "workStartTime": "09:00",
    "workEndTime": "18:00",
    "workDays": [1, 2, 3, 4, 5],
    "requestDelayMin": 5000,
    "requestDelayMax": 15000,
    "detailDwellMin": 10000,
    "detailDwellMax": 30000,
    "actionDwellMin": 8000,
    "actionDwellMax": 18000,
    "scrollMode": "mixed",
    "dailyLimit": 20,
    "hourlyLimit": 6,
    "maxCandidatesPerRun": 5,
    "browseProbability": 0.55,
    "longBreakEvery": 3,
    "longBreakMin": 60000,
    "longBreakMax": 150000,
    "interactionModes": {
        "manualPage": 40,
        "detailClick": 35,
        "filterReview": 25,
    },
    "searchKeywordPool": [
        "Java开发 北京 25-35K",
        "前端工程师 React 上海",
        "Python 后端 杭州",
        "测试开发 深圳 20-30K",
    ],
}

DEFAULT_LLM_CONFIG: dict[str, Any] = {
    "llmEnabled": False,
    "llmProvider": "siliconflow",
    "llmProtocol": "openai-chat",
    "llmApiBase": "https://api.siliconflow.cn/v1",
    "llmModel": "deepseek-ai/DeepSeek-V3.2",
    "llmTemperature": 0.2,
    "llmMaxContextItems": 12,
    "llmMaxTokens": 400,
    "llmTimeoutSeconds": 60,
}

LLM_PROVIDER_PRESETS: dict[str, dict[str, str]] = {
    "openai-codex": {
        "label": "Codex GPT-5.2",
        "protocol": "openai-responses",
        "apiBase": "https://api.openai.com/v1",
        "model": "gpt-5.2-codex",
        "keyUrl": "https://platform.openai.com/api-keys",
    },
    "openai": {
        "label": "OpenAI",
        "protocol": "openai-chat",
        "apiBase": "",
        "model": "",
        "keyUrl": "https://platform.openai.com/api-keys",
    },
    "claude": {
        "label": "Claude",
        "protocol": "anthropic-messages",
        "apiBase": "https://api.anthropic.com/v1",
        "model": "",
        "keyUrl": "https://console.anthropic.com/settings/keys",
    },
    "qwen": {
        "label": "通义千问",
        "protocol": "openai-chat",
        "apiBase": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "model": "",
    },
    "aliyun": {
        "label": "阿里云百炼",
        "protocol": "openai-chat",
        "apiBase": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "model": "",
    },
    "siliconflow": {
        "label": "硅基流动",
        "protocol": "openai-chat",
        "apiBase": "https://api.siliconflow.cn/v1",
        "model": "deepseek-ai/DeepSeek-V3.2",
        "keyUrl": "https://cloud.siliconflow.cn/me/account/ak",
    },
    "zhipu": {
        "label": "清华系智谱 GLM",
        "protocol": "openai-chat",
        "apiBase": "https://open.bigmodel.cn/api/paas/v4",
        "model": "glm-4.5v",
        "keyUrl": "https://bigmodel.cn/usercenter/proj-mgmt/apikeys",
    },
    "deepseek": {
        "label": "DeepSeek",
        "protocol": "openai-chat",
        "apiBase": "https://api.deepseek.com/v1",
        "model": "",
    },
    "custom": {
        "label": "自定义 OpenAI-compatible",
        "protocol": "openai-chat",
        "apiBase": "",
        "model": "",
    },
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def date_str(offset_days: int = 0) -> str:
    return (datetime.now() + timedelta(days=offset_days)).date().isoformat()


def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_columns(conn: sqlite3.Connection, table: str, columns: dict[str, str]) -> None:
    existing = {
        str(row["name"])
        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }
    for name, definition in columns.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS candidates (
                id TEXT PRIMARY KEY,
                name TEXT,
                role TEXT,
                source TEXT,
                account_name TEXT,
                account_platform TEXT,
                education TEXT,
                experience TEXT,
                expected_salary TEXT,
                score INTEGER,
                recommendation TEXT,
                received_date TEXT,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS recommendations (
                id TEXT PRIMARY KEY,
                candidate_id TEXT,
                name TEXT,
                role TEXT,
                source TEXT,
                account_name TEXT,
                score INTEGER,
                recommendation TEXT,
                next_step TEXT,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reports (
                id TEXT PRIMARY KEY,
                candidate_id TEXT,
                name TEXT,
                role TEXT,
                report TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS agent_conversations (
                id TEXT PRIMARY KEY,
                channel TEXT NOT NULL,
                sender TEXT,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                raw_json TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS job_requirements (
                id TEXT PRIMARY KEY,
                role TEXT NOT NULL,
                normalized_role TEXT NOT NULL UNIQUE,
                source TEXT,
                account_name TEXT,
                requirement TEXT NOT NULL,
                status TEXT DEFAULT 'active',
                failure_reason TEXT DEFAULT '',
                source_url TEXT,
                raw_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )
        ensure_columns(conn, "job_requirements", {
            "status": "TEXT DEFAULT 'active'",
            "failure_reason": "TEXT DEFAULT ''",
        })
        cleanup_job_requirements(conn)
        cleanup_candidate_names(conn)


def normalize_host_header(value: str) -> str:
    host = str(value or "").strip().lower()
    if not host:
        return ""
    if host.startswith("[") and "]" in host:
        return host[1:host.index("]")]
    return host.split(":", 1)[0]


def is_allowed_host(handler: SimpleHTTPRequestHandler) -> bool:
    if not ALLOWED_HOSTS:
        return True
    host = normalize_host_header(handler.headers.get("Host", ""))
    return not host or host in ALLOWED_HOSTS


def is_allowed_origin_value(origin: str) -> bool:
    origin = str(origin or "").strip().rstrip("/")
    if not origin:
        return True
    if origin.startswith("chrome-extension://"):
        return True
    if not ALLOWED_ORIGINS:
        return True
    return origin in ALLOWED_ORIGINS


def cors_origin(handler: SimpleHTTPRequestHandler) -> str:
    origin = str(handler.headers.get("Origin") or "").strip().rstrip("/")
    if origin and is_allowed_origin_value(origin):
        return origin
    return ALLOWED_ORIGINS[0] if ALLOWED_ORIGINS else "*"


def client_ip(handler: SimpleHTTPRequestHandler) -> str:
    return handler.client_address[0] if handler.client_address else ""


def is_allowed_client_ip(handler: SimpleHTTPRequestHandler) -> bool:
    if not ALLOWED_CLIENT_IPS:
        return True
    ip_text = client_ip(handler)
    try:
        ip = ipaddress.ip_address(ip_text)
    except ValueError:
        return False
    for rule in ALLOWED_CLIENT_IPS:
        try:
            if "/" in rule and ip in ipaddress.ip_network(rule, strict=False):
                return True
            if ip == ipaddress.ip_address(rule):
                return True
        except ValueError:
            continue
    return False


def is_rate_limited(handler: SimpleHTTPRequestHandler) -> bool:
    if RATE_LIMIT_PER_MINUTE <= 0:
        return False
    parsed = urllib.parse.urlparse(handler.path)
    if not parsed.path.startswith("/api/"):
        return False
    key = (client_ip(handler), parsed.path)
    now = time.time()
    with RATE_LOCK:
        bucket = [stamp for stamp in RATE_BUCKETS.get(key, []) if now - stamp < 60]
        if len(bucket) >= RATE_LIMIT_PER_MINUTE:
            RATE_BUCKETS[key] = bucket
            return True
        bucket.append(now)
        RATE_BUCKETS[key] = bucket
    return False


def write_common_headers(handler: SimpleHTTPRequestHandler) -> None:
    setattr(handler, "_common_headers_written", True)
    handler.send_header("Access-Control-Allow-Origin", cors_origin(handler))
    handler.send_header("Access-Control-Allow-Headers", "Content-Type, X-Admin-Token")
    handler.send_header("Access-Control-Allow-Methods", "GET,HEAD,POST,OPTIONS")
    handler.send_header("X-Content-Type-Options", "nosniff")
    handler.send_header("X-Frame-Options", "DENY")
    handler.send_header("Referrer-Policy", "same-origin")
    handler.send_header("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    handler.send_header("Cache-Control", "no-store")


def guard_request(handler: SimpleHTTPRequestHandler) -> bool:
    origin = str(handler.headers.get("Origin") or "")
    if not is_allowed_host(handler):
        json_response(handler, {"success": False, "message": "Host 不在白名单内"}, 403)
        return False
    if not is_allowed_origin_value(origin):
        json_response(handler, {"success": False, "message": "Origin 不在白名单内"}, 403)
        return False
    if not is_allowed_client_ip(handler):
        json_response(handler, {"success": False, "message": "客户端 IP 不在白名单内"}, 403)
        return False
    if is_rate_limited(handler):
        json_response(handler, {"success": False, "message": "请求过于频繁，请稍后再试"}, 429)
        return False
    if ADMIN_TOKEN and handler.command in {"POST", "PUT", "PATCH", "DELETE"}:
        parsed = urllib.parse.urlparse(handler.path)
        public_paths = {"/api/dingtalk/callback", "/api/dingtalk/callback-test"}
        if parsed.path not in public_paths and handler.headers.get("X-Admin-Token", "") != ADMIN_TOKEN:
            json_response(handler, {"success": False, "message": "缺少管理令牌"}, 401)
            return False
    return True


def current_security_request(handler: SimpleHTTPRequestHandler) -> dict[str, str]:
    origin = normalize_origin_value(handler.headers.get("Origin"))
    return {
        "host": normalize_host_header(handler.headers.get("Host", "")),
        "origin": origin or request_base_url(handler),
        "clientIp": client_ip(handler),
    }


def get_security_allowlist_config(handler: SimpleHTTPRequestHandler) -> dict[str, Any]:
    return {
        "success": True,
        "filePath": str(SECURITY_ALLOWLIST_PATH),
        "editable": {
            "allowedHosts": clean_allowed_hosts(SECURITY_ALLOWLIST.get("allowedHosts", [])),
            "allowedOrigins": clean_allowed_origins(SECURITY_ALLOWLIST.get("allowedOrigins", [])),
            "allowedClientIps": clean_allowed_client_ips(SECURITY_ALLOWLIST.get("allowedClientIps", [])),
            "rateLimitPerMinute": int(SECURITY_ALLOWLIST.get("rateLimitPerMinute") or RATE_LIMIT_PER_MINUTE),
            "maxJsonBodyBytes": int(SECURITY_ALLOWLIST.get("maxJsonBodyBytes") or MAX_JSON_BODY_BYTES),
        },
        "effective": {
            "allowedHosts": ALLOWED_HOSTS,
            "allowedOrigins": ALLOWED_ORIGINS,
            "allowedClientIps": ALLOWED_CLIENT_IPS,
            "clientIpAllowlistEnabled": bool(ALLOWED_CLIENT_IPS),
            "rateLimitPerMinute": RATE_LIMIT_PER_MINUTE,
            "maxJsonBodyBytes": MAX_JSON_BODY_BYTES,
        },
        "current": current_security_request(handler),
        "env": {
            "allowedHosts": _split_env_list("RECRUITMENT_ALLOWED_HOSTS"),
            "allowedOrigins": _split_env_list("RECRUITMENT_ALLOWED_ORIGINS"),
            "allowedClientIps": [
                *_split_env_list("RECRUITMENT_ALLOWED_CLIENT_IPS"),
                *_split_env_list("RECRUITMENT_IP_ALLOWLIST"),
            ],
        },
    }


def save_security_allowlist_config(payload: dict[str, Any], handler: SimpleHTTPRequestHandler) -> dict[str, Any]:
    try:
        hosts = clean_allowed_hosts(payload.get("allowedHosts", []))
        origins = clean_allowed_origins(payload.get("allowedOrigins", []))
        client_ips = clean_allowed_client_ips(payload.get("allowedClientIps", []))
    except ValueError as exc:
        return {"success": False, "message": str(exc)}

    if payload.get("keepCurrentAccess", True):
        current = current_security_request(handler)
        current_host = normalize_host_header(current.get("host", ""))
        current_origin = normalize_origin_value(current.get("origin", ""))
        if current_host and current_host not in hosts:
            hosts.append(current_host)
        if current_origin and current_origin not in origins:
            origins.append(current_origin)

    try:
        rate_limit = int(payload.get("rateLimitPerMinute") or RATE_LIMIT_PER_MINUTE or 0)
    except (TypeError, ValueError):
        rate_limit = RATE_LIMIT_PER_MINUTE
    try:
        max_json = int(payload.get("maxJsonBodyBytes") or MAX_JSON_BODY_BYTES or 1048576)
    except (TypeError, ValueError):
        max_json = MAX_JSON_BODY_BYTES

    next_config = {
        "allowedHosts": sorted(set(hosts)),
        "allowedOrigins": sorted(set(origins)),
        "allowedClientIps": sorted(set(client_ips)),
        "rateLimitPerMinute": max(0, min(rate_limit, 10000)),
        "maxJsonBodyBytes": max(1024, min(max_json, 20 * 1024 * 1024)),
    }
    SECURITY_ALLOWLIST_PATH.parent.mkdir(parents=True, exist_ok=True)
    SECURITY_ALLOWLIST_PATH.write_text(json.dumps(next_config, ensure_ascii=False, indent=2) + "\n", "utf-8")
    reload_security_runtime(next_config)
    return get_security_allowlist_config(handler)


def json_response(handler: SimpleHTTPRequestHandler, payload: Any, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    write_common_headers(handler)
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(
    handler: SimpleHTTPRequestHandler,
    body: str,
    status: int = 200,
    content_type: str = "text/plain; charset=utf-8",
    filename: str | None = None,
) -> None:
    data = body.encode("utf-8-sig")
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    write_common_headers(handler)
    if filename:
        handler.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def binary_response(
    handler: SimpleHTTPRequestHandler,
    data: bytes,
    content_type: str,
    filename: str | None = None,
    head_only: bool = False,
) -> None:
    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    write_common_headers(handler)
    if filename:
        quoted = urllib.parse.quote(filename)
        handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quoted}")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    if not head_only:
        handler.wfile.write(data)


def read_json(handler: SimpleHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if length > MAX_BODY_BYTES:
        raise ValueError(f"request body too large, max {MAX_BODY_BYTES} bytes")
    if length <= 0:
        return {}
    if length > MAX_JSON_BODY_BYTES:
        raise ValueError(f"请求体过大，最大允许 {MAX_JSON_BODY_BYTES} 字节")
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


def normalize_candidate_payload(candidate: dict[str, Any]) -> dict[str, Any]:
    cleaned = dict(candidate or {})
    name = str(cleaned.get("name") or "").strip()
    top_level_text = sanitize_candidate_snapshot(cleaned.get("topLevelText") or cleaned.get("rawText") or "")
    if top_level_text:
        cleaned["topLevelText"] = top_level_text
        cleaned["rawText"] = top_level_text
    raw_text = " ".join(
        str(cleaned.get(key) or "")
        for key in ("rawText", "summary", "jobRequirement")
    )
    if is_generic_candidate_name(name):
        name = ""
    if not name:
        name = extract_candidate_name_from_text(" ".join(
            str(cleaned.get(key) or "")
            for key in ("topLevelText", "rawText", "summary")
        ))
    if not name:
        name = unresolved_candidate_display_name(cleaned)
        cleaned["_nameUnresolved"] = True
    cleaned["name"] = name
    polluted = looks_like_polluted_candidate(cleaned, raw_text)
    if polluted:
        evaluation = dict(cleaned.get("evaluation") or {})
        evaluation["score"] = 0
        evaluation["recommendation"] = "待复核"
        evaluation["nextStep"] = "数据采集范围异常，请重新打开候选人详情后采集"
        cleaned["evaluation"] = evaluation
        cleaned["score"] = 0
        cleaned["recommendation"] = "待复核"
        cleaned["_invalidRecommendation"] = True
        cleaned["_invalidReason"] = "候选人详情采集混入列表或整页文本，已阻止进入推荐报告"
    return cleaned


def validate_candidate_minimum(candidate: dict[str, Any]) -> tuple[bool, str]:
    name = str(candidate.get("name") or "").strip()
    role = str(candidate.get("role") or candidate.get("jobRole") or "").strip()
    education = str(candidate.get("education") or "").strip()
    experience = str(candidate.get("experience") or "").strip()
    if candidate.get("_nameUnresolved") or is_generic_candidate_name(name):
        return False, "候选人姓名未准确识别"
    if not name:
        return False, "缺少候选人姓名"
    if not role:
        return False, "缺少沟通职位"
    if not (education or experience):
        return False, "缺少学历或经验信息"
    return True, ""


def sanitize_candidate_snapshot(text: Any) -> str:
    normalized = " ".join(str(text or "").split()).strip()
    if not normalized:
        return ""
    stop_markers = [
        "工作经历", "项目经历", "教育经历", "资格证书", "求职期望",
        "沟通记录", "聊天记录", "全部职位", "新招呼", "沟通中",
        "账号权益", "招聘规范", "职位管理", "推荐牛人", "批量",
    ]
    for marker in stop_markers:
        index = normalized.find(marker)
        if index > 8:
            normalized = normalized[:index].strip()
    return normalized[:900]


def is_generic_candidate_name(name: str) -> bool:
    text = str(name or "").strip()
    return text in {"候选人", "牛人", "求职者", "用户", "先生", "女士", "姓名", "未识别", "未识别姓名", "未识别候选人"} or text.startswith("未识别候选人-")


def unresolved_candidate_display_name(candidate: dict[str, Any]) -> str:
    role = str(candidate.get("role") or candidate.get("jobRole") or "候选人").strip()
    role = re.sub(r"\s+", "", role)[:18] or "候选人"
    raw_id = str(candidate.get("id") or candidate.get("candidate_id") or candidate.get("receivedTime") or now_iso())
    suffix = hashlib.sha1(raw_id.encode()).hexdigest()[:6]
    return f"未识别候选人-{role}-{suffix}"


def is_likely_candidate_name(name: str) -> bool:
    text = re.sub(r"\s+", "", str(name or "").strip())
    if is_generic_candidate_name(text):
        return False
    if not re.fullmatch(r"[\u4e00-\u9fa5]{2,4}", text):
        return False
    blocked = (
        "工程师|开发|分析|运营|产品|经理|实习|岗位|职位|简历|数据|算法|测试|前端|后端|"
        "电气|结构|工艺|平台|系统|智能|求职|招聘|在线|离线|本科|硕士|博士|大专|"
        "经验|工作|公司|项目|沟通|您好|你好|谢谢|匹配|期待"
    )
    return not re.search(blocked, text)


def clean_candidate_name(value: Any) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return ""
    text = re.sub(r"^[红绿]点", "", text)
    text = re.sub(r"^\d+\s*", "", text)
    text = re.sub(r"^\d{1,2}:\d{2}\s*", "", text)
    text = re.sub(r"[（(].*?[）)]", "", text)
    text = re.sub(r"\s*(在线|离线|活跃|已读|未读).*$", "", text)
    text = re.sub(r"\s*(沟通职位|期望|应聘|投递).*$", "", text).strip()
    match = re.match(r"[\u4e00-\u9fa5]{2,4}", text)
    name = match.group(0) if match else text
    return name if is_likely_candidate_name(name) else ""


def extract_candidate_name_from_text(text: Any) -> str:
    normalized = " ".join(str(text or "").split()).strip()
    if not normalized:
        return ""
    compact = re.sub(r"^[\s\d红绿点未读已读]+", "", normalized)
    compact = re.sub(r"^\d{1,2}:\d{2}\s*", "", compact)
    compact = re.sub(r"^牛人分析器\s*", "", compact).strip()
    patterns = [
        r"^([\u4e00-\u9fa5]{2,4})(?=\s*(?:[·•]|在线|离线|\d{2}岁|男|女|本科|大专|硕士|博士))",
        r"^([\u4e00-\u9fa5]{2,4})(?=\s+(?:前端|后端|测试|数据|运营|产品|结构|电气|工艺|算法|嵌入式|Java|C\+\+|DevOps|SQE|PLM|ERP|具身|行业|动力|底盘|机器人))",
        r"^([\u4e00-\u9fa5]{2,4})(?=\s+[\u4e00-\u9fa5A-Za-z/+#（）()]+(?:工程师|开发|分析|运营|产品|实习|经理|算法))",
        r"(?:^|\s)([\u4e00-\u9fa5]{2,4})(?=\s*(?:[·•]|在线|离线|\d{2}岁))",
        r"(?:姓名|候选人)[:：]\s*([\u4e00-\u9fa5]{2,4})",
    ]
    for source in (compact, normalized):
        for pattern in patterns:
            match = re.search(pattern, source, flags=re.I)
            name = clean_candidate_name(match.group(1) if match else "")
            if name:
                return name
    return ""


def looks_like_polluted_candidate(candidate: dict[str, Any], raw_text: str) -> bool:
    name = str(candidate.get("name") or "").strip()
    role = str(candidate.get("role") or "")
    candidate_id = str(candidate.get("id") or candidate.get("candidate_id") or "")
    text = raw_text or ""
    unresolved_name = bool(candidate.get("_nameUnresolved")) or is_generic_candidate_name(name)
    if not text:
        return is_generic_candidate_name(name) and candidate_id.startswith("chat_")
    page_markers = sum(1 for marker in ["全部职位", "新招呼", "沟通中", "账号权益", "招聘规范", "职位管理"] if marker in text)
    repeated_roles = len(set(re.findall(r"[\u4e00-\u9fa5A-Za-z/+-]+(?:工程师|分析|开发|产品|运营)[^\s，。|]{0,18}\(J\d+\)", text)))
    if unresolved_name and (page_markers >= 2 or repeated_roles >= 3):
        return True
    if unresolved_name and candidate_id.startswith("chat_"):
        return True
    if role and repeated_roles >= 6 and page_markers >= 2:
        return True
    return False


def upsert_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    candidate = normalize_candidate_payload(candidate)
    ok, reason = validate_candidate_minimum(candidate)
    if not ok:
        return {
            "success": False,
            "message": f"候选人未入库：{reason}",
            "quality": {"ok": False, "reason": reason},
        }
    candidate_id = str(candidate.get("id") or hashlib.sha1(json.dumps(candidate, ensure_ascii=False).encode()).hexdigest())
    evaluation = candidate.get("evaluation") or {}
    with connect() as conn:
        existing = conn.execute("SELECT id FROM candidates WHERE id = ?", (candidate_id,)).fetchone()
        created_at = now_iso()
        if existing:
            created_at = conn.execute("SELECT created_at FROM candidates WHERE id = ?", (candidate_id,)).fetchone()["created_at"]
        conn.execute(
            """
            INSERT OR REPLACE INTO candidates
            (id, name, role, source, account_name, account_platform, education, experience,
             expected_salary, score, recommendation, received_date, raw_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                candidate_id,
                candidate.get("name", ""),
                candidate.get("role", ""),
                candidate.get("source", ""),
                candidate.get("accountName", ""),
                candidate.get("accountPlatform", ""),
                candidate.get("education", ""),
                candidate.get("experience", ""),
                candidate.get("expectedSalary", ""),
                int(evaluation.get("score") or 0),
                evaluation.get("recommendation", ""),
                candidate.get("receivedDate", date_str()),
                json.dumps(candidate, ensure_ascii=False),
                created_at,
                now_iso(),
            ),
        )
    return {"success": True, "id": candidate_id}


def normalize_role(role: str) -> str:
    return "".join(str(role or "").lower().split())


def normalize_account_name(account_name: Any) -> str:
    return re.sub(r"\s+", "", str(account_name or "").strip().lower())


def job_requirement_role_key(role: str) -> str:
    return normalize_role(role)


def account_label(account_name: Any) -> str:
    return str(account_name or "").strip() or "未识别"


def normalize_managed_account(value: Any) -> dict[str, Any] | None:
    if isinstance(value, dict):
        name = str(value.get("name") or value.get("accountName") or value.get("account_name") or "").strip()
        platform = str(value.get("platform") or value.get("accountPlatform") or "BOSS直聘").strip() or "BOSS直聘"
        source = str(value.get("source") or "manual").strip() or "manual"
    else:
        name = str(value or "").strip()
        platform = "BOSS直聘"
        source = "manual"
    if not name:
        return None
    return {
        "name": name,
        "platform": platform,
        "source": source,
        "updatedAt": now_iso(),
    }


def managed_accounts_from_settings(settings: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    settings = settings or get_settings()
    by_key: dict[str, dict[str, Any]] = {}

    def add(value: Any, default_source: str = "manual") -> None:
        item = normalize_managed_account(value)
        if not item:
            return
        if default_source and item.get("source") == "manual":
            item["source"] = default_source
        key = normalize_account_name(item["name"])
        if key and key not in by_key:
            by_key[key] = item

    for item in settings.get("managedAccounts") or []:
        add(item, "manual")
    add({
        "name": settings.get("accountName"),
        "platform": settings.get("accountPlatform") or "BOSS直聘",
        "source": "current",
    }, "current")
    detected = settings.get("detectedAccount") if isinstance(settings.get("detectedAccount"), dict) else {}
    add({
        "name": detected.get("name"),
        "platform": detected.get("platform") or settings.get("accountPlatform") or "BOSS直聘",
        "source": detected.get("source") or "detected",
    }, "detected")
    return sorted(by_key.values(), key=lambda item: item["name"])


def save_managed_account(payload: dict[str, Any]) -> dict[str, Any]:
    item = normalize_managed_account({
        "name": payload.get("name") or payload.get("accountName"),
        "platform": payload.get("platform") or payload.get("accountPlatform"),
        "source": "manual",
    })
    if not item:
        return {"success": False, "message": "账号名称不能为空"}
    settings = get_settings()
    accounts = managed_accounts_from_settings(settings)
    by_key = {normalize_account_name(account["name"]): account for account in accounts}
    by_key[normalize_account_name(item["name"])] = item
    save_settings({
        "managedAccounts": sorted(by_key.values(), key=lambda account: account["name"]),
        "accountName": item["name"],
        "accountPlatform": item["platform"],
        "accountNameManual": True,
    })
    return {"success": True, "account": item, "items": get_accounts()}


def delete_managed_account(payload: dict[str, Any]) -> dict[str, Any]:
    name = str(payload.get("name") or payload.get("accountName") or "").strip()
    if not name:
        return {"success": False, "message": "账号名称不能为空"}
    target = normalize_account_name(name)
    settings = get_settings()
    accounts = [
        account for account in managed_accounts_from_settings(settings)
        if normalize_account_name(account["name"]) != target
    ]
    updates: dict[str, Any] = {"managedAccounts": accounts}
    if normalize_account_name(settings.get("accountName")) == target:
        updates.update({
            "accountName": accounts[0]["name"] if accounts else "",
            "accountPlatform": accounts[0]["platform"] if accounts else "BOSS直聘",
            "accountNameManual": bool(accounts),
        })
    save_settings(updates)
    return {"success": True, "items": get_accounts()}


def scoped_role_key(role: str, account_name: Any = "") -> str:
    return job_requirement_role_key(role)


def sanitize_job_requirement_text(requirement: str) -> str:
    text = re.sub(r"\s+", " ", str(requirement or "")).strip()
    if not text:
        return ""
    text = re.sub(
        r"(工作内容|工作职责|岗位职责|职位职责|职位描述|职位介绍|工作要求|任职要求|任职资格|岗位要求|职位要求)",
        r" \1 ",
        text,
    )
    has_resume_markers = bool(re.search(r"工作经历|项目经验|项目经历|教育经历|期望职位|求职期望", text))
    has_job_detail_markers = bool(re.search(r"职位详情|职位描述|职位职责|任职要求|任职资格|岗位要求|职位要求|薪资详情|工作地址|职位发布", text))
    if has_resume_markers and not has_job_detail_markers:
        return ""
    stop_markers = [
        "薪资详情", "职位福利", "工作地点", "工作地址", "职位发布", "公司介绍", "工商信息", "竞争力分析",
        "相似职位", "推荐职位", "沟通职位", "立即沟通", "在线沟通", "全部职位",
        "新招呼", "沟通中", "账号权益", "招聘规范", "我的客服", "招聘数据",
    ]

    def section(start_titles: list[str], stop_titles: list[str]) -> str:
        start_index = -1
        title = ""
        for candidate_title in start_titles:
            index = text.find(candidate_title)
            if index >= 0:
                start_index = index
                title = candidate_title
                break
        if start_index < 0:
            return ""
        body_start = start_index + len(title)
        stops = [(text.find(marker, body_start), marker) for marker in stop_titles if text.find(marker, body_start) >= 0]
        end = sorted(stops, key=lambda item: item[0])[0][0] if stops else min(len(text), body_start + 2600)
        body = re.sub(r"^[：:，,\s-]+", "", text[body_start:end]).strip()
        if len(body) < 20:
            return ""
        return f"{title}：{body[:3000]}"

    content = section(
        ["工作内容", "工作职责", "岗位职责", "职位职责", "职位描述", "职位介绍"],
        ["工作要求", "任职要求", "任职资格", "岗位要求", "职位要求", *stop_markers],
    )
    requirement_text = section(
        ["工作要求", "任职要求", "任职资格", "职位要求", "岗位要求"],
        stop_markers,
    )
    return "\n".join(part for part in (content, requirement_text) if part).strip()[:12000]


def is_valid_job_requirement_text(requirement: str) -> bool:
    text = re.sub(r"\s+", " ", str(requirement or "")).strip()
    if len(text) < 20:
        return False
    has_job_section = bool(re.search(r"工作内容|工作职责|岗位职责|职位职责|职位描述|职位介绍|工作要求|任职要求|任职资格|岗位要求|职位要求", text))
    polluted_markers = [
        "新招呼", "沟通中", "全部职位", "账号权益", "招聘规范", "BOSS您好", "Boss，您好",
        "您好，我叫", "您好，我是", "您好！我是", "你好，我是", "进一步沟通",
        "期待进一步", "我的简历", "详细简历", "完全匹配", "挺适合",
        "对贵公司", "对贵岗位", "方便发一份您的简历",
    ]
    if any(marker in text for marker in polluted_markers):
        return False
    repeated_roles = set(re.findall(r"[\u4e00-\u9fa5A-Za-z/+-]+(?:工程师|分析|开发|产品|运营)[^\s，。|]{0,18}\(J\d+\)", text))
    if len(repeated_roles) >= 2:
        return False
    has_content = bool(re.search(r"工作内容|工作职责|岗位职责|职位职责|职位描述|职位介绍", text))
    has_requirement = bool(re.search(r"工作要求|任职要求|任职资格|岗位要求|职位要求", text))
    return has_job_section and (has_content or has_requirement)


def job_requirement_is_active(row: dict[str, Any] | sqlite3.Row | None) -> bool:
    if not row:
        return False
    item = dict(row)
    return str(item.get("status") or "active") == "active" and is_valid_job_requirement_text(str(item.get("requirement") or ""))


def upsert_job_requirement(payload: dict[str, Any]) -> dict[str, Any]:
    role = str(payload.get("role") or "").strip()
    requirement = sanitize_job_requirement_text(payload.get("requirement") or payload.get("jobRequirement") or "")
    account_name = str(payload.get("accountName") or payload.get("account_name") or "").strip()
    if not role:
        return {"success": False, "message": "岗位名称不能为空"}
    if not is_valid_job_requirement_text(requirement):
        return {"success": False, "message": "岗位要求内容不符合工作内容/工作要求格式，未保存"}
    normalized = scoped_role_key(role, account_name)
    item_id = hashlib.sha1(normalized.encode()).hexdigest()
    now = now_iso()
    with connect() as conn:
        existing = conn.execute(
            "SELECT created_at FROM job_requirements WHERE normalized_role = ?",
            (normalized,),
        ).fetchone()
        conn.execute(
            """
            INSERT OR REPLACE INTO job_requirements
            (id, role, normalized_role, source, account_name, requirement, status, failure_reason, source_url, raw_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item_id,
                role,
                normalized,
                payload.get("source", ""),
                account_name,
                requirement[:12000],
                "active",
                "",
                payload.get("sourceUrl", ""),
                json.dumps(payload, ensure_ascii=False),
                existing["created_at"] if existing else now,
                now,
            ),
        )
    matched = match_candidates_with_job_requirements(role, account_name)
    return {"success": True, "id": item_id, "matchedCandidates": matched["updated"]}


def upsert_pending_job_requirement(payload: dict[str, Any]) -> dict[str, Any]:
    role = str(payload.get("role") or "").strip()
    account_name = str(payload.get("accountName") or payload.get("account_name") or "").strip()
    if not role:
        return {"success": False, "message": "岗位名称不能为空"}
    normalized = scoped_role_key(role, account_name)
    item_id = hashlib.sha1(normalized.encode()).hexdigest()
    now = now_iso()
    reason = str(payload.get("reason") or payload.get("failureReason") or "插件未能识别工作内容/工作要求，请手动补录").strip()
    with connect() as conn:
        existing = conn.execute(
            "SELECT * FROM job_requirements WHERE normalized_role = ?",
            (normalized,),
        ).fetchone()
        if job_requirement_is_active(existing):
            return {"success": True, "id": dict(existing)["id"], "status": "active", "message": "岗位要求已存在"}
        conn.execute(
            """
            INSERT OR REPLACE INTO job_requirements
            (id, role, normalized_role, source, account_name, requirement, status, failure_reason, source_url, raw_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item_id,
                role,
                normalized,
                payload.get("source", ""),
                account_name,
                "",
                "pending",
                reason[:500],
                payload.get("sourceUrl", ""),
                json.dumps(payload, ensure_ascii=False),
                (dict(existing).get("created_at") if existing else None) or now,
                now,
            ),
        )
    return {"success": True, "id": item_id, "status": "pending", "message": "岗位已标记为待补录"}


def cleanup_job_requirements(conn: sqlite3.Connection) -> None:
    rows = [dict(row) for row in conn.execute("SELECT * FROM job_requirements").fetchall()]
    if not rows:
        return
    by_role: dict[str, dict[str, Any]] = {}
    for row in rows:
        role = str(row.get("role") or "").strip()
        requirement = sanitize_job_requirement_text(row.get("requirement") or "")
        status = str(row.get("status") or "active")
        if not role:
            continue
        key = job_requirement_role_key(role)
        if status == "pending" and not is_valid_job_requirement_text(requirement):
            existing = by_role.get(key)
            if not existing:
                by_role[key] = {
                    **row,
                    "id": hashlib.sha1(key.encode()).hexdigest(),
                    "normalized_role": key,
                    "requirement": "",
                    "status": "pending",
                    "failure_reason": row.get("failure_reason") or "待手动补录",
                }
            continue
        if not is_valid_job_requirement_text(requirement):
            continue
        existing = by_role.get(key)
        if (
            not existing
            or str(existing.get("status") or "active") == "pending"
            or len(requirement) > len(str(existing.get("requirement") or ""))
            or str(row.get("updated_at") or "") > str(existing.get("updated_at") or "")
        ):
            by_role[key] = {
                **row,
                "id": hashlib.sha1(key.encode()).hexdigest(),
                "normalized_role": key,
                "requirement": requirement[:12000],
                "status": "active",
                "failure_reason": "",
            }

    conn.execute("DELETE FROM job_requirements")
    for item in by_role.values():
        conn.execute(
            """
            INSERT OR REPLACE INTO job_requirements
            (id, role, normalized_role, source, account_name, requirement, status, failure_reason, source_url, raw_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.get("id"),
                item.get("role", ""),
                item.get("normalized_role", ""),
                item.get("source", ""),
                item.get("account_name", ""),
                item.get("requirement", ""),
                item.get("status", "active"),
                item.get("failure_reason", ""),
                item.get("source_url", ""),
                item.get("raw_json", ""),
                item.get("created_at") or now_iso(),
                item.get("updated_at") or now_iso(),
            ),
        )


def cleanup_candidate_names(conn: sqlite3.Connection) -> None:
    rows = [dict(row) for row in conn.execute(
        "SELECT id, name, raw_json FROM candidates WHERE COALESCE(name, '') = '' OR name IN ('候选人','牛人','求职者','用户','先生','女士','未识别','未识别姓名')"
    ).fetchall()]
    for row in rows:
        try:
            raw = json.loads(row.get("raw_json") or "{}")
        except json.JSONDecodeError:
            raw = {}
        name = extract_candidate_name_from_text(" ".join(
            str(raw.get(key) or "")
            for key in ("topLevelText", "rawText", "summary")
        ))
        if not name and is_generic_candidate_name(row.get("name", "")):
            name = ""
        if not name:
            name = unresolved_candidate_display_name({
                **raw,
                "id": row.get("id"),
                "role": raw.get("role") or "",
            })
            raw["_nameUnresolved"] = True
        if name == (row.get("name") or ""):
            continue
        raw["name"] = name
        conn.execute(
            "UPDATE candidates SET name = ?, raw_json = ?, updated_at = ? WHERE id = ?",
            (name, json.dumps(raw, ensure_ascii=False), now_iso(), row["id"]),
        )


def get_job_requirement(role: str, account_name: str = "") -> dict[str, Any] | None:
    normalized = normalize_role(role)
    if not normalized:
        return None
    scoped_normalized = scoped_role_key(role, account_name)
    with connect() as conn:
        if account_name:
            row = conn.execute(
                "SELECT * FROM job_requirements WHERE normalized_role = ?",
                (scoped_normalized,),
            ).fetchone()
            if job_requirement_is_active(row):
                return dict(row)
            loose_account = conn.execute(
                """
                SELECT * FROM job_requirements
                WHERE COALESCE(NULLIF(account_name, ''), '未识别') = ?
                  AND (normalized_role LIKE ? OR ? LIKE '%' || normalized_role || '%' OR normalized_role LIKE ?)
                ORDER BY updated_at DESC LIMIT 1
                """,
                (account_label(account_name), f"%{normalized}%", normalized, f"%::{normalized}"),
            ).fetchone()
            if job_requirement_is_active(loose_account):
                return dict(loose_account)
        row = conn.execute(
            "SELECT * FROM job_requirements WHERE normalized_role = ?",
            (normalized,),
        ).fetchone()
        if job_requirement_is_active(row):
            return dict(row)
        loose = conn.execute(
            "SELECT * FROM job_requirements WHERE normalized_role LIKE ? OR ? LIKE '%' || normalized_role || '%' ORDER BY updated_at DESC LIMIT 1",
            (f"%{normalized}%", normalized),
        ).fetchone()
        return dict(loose) if job_requirement_is_active(loose) else None


def list_job_requirements(limit: int = 200, filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
    return list_rows("job_requirements", limit, filters=filters)


def extract_requirement_keywords_backend(text: str) -> list[str]:
    normalized = str(text or "").lower()
    keywords = [
        "java", "spring", "springboot", "mysql", "redis", "python", "django", "flask",
        "go", "golang", "react", "vue", "typescript", "javascript", "node", "测试",
        "自动化", "selenium", "playwright", "性能", "运维", "kubernetes", "docker",
        "算法", "数据", "产品", "项目管理", "招聘", "销售", "客服", "运营",
    ]
    return [item for item in keywords if item in normalized]


def recommendation_label_by_score(score: int) -> str:
    if score >= 80:
        return "强烈推荐"
    if score >= 60:
        return "非常推荐"
    if score >= 40:
        return "推荐"
    return "不推荐"


def match_candidate_with_requirement(candidate: dict[str, Any], job: dict[str, Any]) -> dict[str, Any]:
    raw: dict[str, Any]
    try:
        raw = json.loads(candidate.get("raw_json") or "{}")
    except Exception:
        raw = {}

    requirement = str(job.get("requirement") or "")
    resume_text = " ".join(
        str(value or "") for value in [
            candidate.get("name"), candidate.get("role"), candidate.get("education"),
            candidate.get("experience"), candidate.get("expected_salary"),
            raw.get("skills"), raw.get("summary"), raw.get("workExperience"), raw.get("projects"),
            raw.get("rawText"), raw.get("description"),
        ]
    ).lower()
    keywords = extract_requirement_keywords_backend(requirement)
    matched = [item for item in keywords if item in resume_text]
    missing = [item for item in keywords if item not in resume_text]
    ratio = 1.0 if not keywords else len(matched) / len(keywords)
    if ratio >= 0.68:
        verdict = "匹配"
        adjustment = 6
    elif ratio >= 0.38:
        verdict = "部分匹配"
        adjustment = 0
    else:
        verdict = "不匹配"
        adjustment = -8

    evaluation = raw.get("evaluation") if isinstance(raw.get("evaluation"), dict) else {}
    dimensions = evaluation.get("dimensions") if isinstance(evaluation.get("dimensions"), dict) else {}
    dimensions["jd"] = {
        "match": verdict,
        "ratio": round(ratio, 2),
        "matched": matched[:12],
        "missing": missing[:12],
        "source": "岗位要求库",
    }
    evaluation["dimensions"] = dimensions
    reasons = evaluation.get("rejectionReasons") if isinstance(evaluation.get("rejectionReasons"), list) else []
    if verdict == "不匹配" and missing:
        reasons = [f"岗位要求缺失：{', '.join(missing[:6])}", *reasons]
    evaluation["rejectionReasons"] = reasons[:12]
    raw["evaluation"] = evaluation
    raw["jobRequirement"] = requirement
    raw["jobRequirementRole"] = job.get("role") or candidate.get("role")

    base_score = int(candidate.get("score") or evaluation.get("score") or 0)
    next_score = max(0, min(100, base_score + adjustment))
    recommendation = recommendation_label_by_score(next_score)
    evaluation["score"] = next_score
    evaluation["recommendation"] = recommendation

    return {
        "raw_json": json.dumps(raw, ensure_ascii=False),
        "score": next_score,
        "recommendation": recommendation,
    }


def match_candidates_with_job_requirements(role: str = "", account: str = "", account_exact: bool = True) -> dict[str, Any]:
    updated = 0
    skipped = 0
    matched_items: list[dict[str, Any]] = []
    with connect() as conn:
        where: list[str] = []
        params: list[Any] = []
        if role:
            where.append("role = ?")
            params.append(role)
        if account:
            operator = "=" if account_exact else "LIKE"
            where.append(f"COALESCE(NULLIF(account_name, ''), '未识别') {operator} ?")
            params.append(account if account_exact else f"%{account}%")
        sql = "SELECT * FROM candidates"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY updated_at DESC"
        candidates = conn.execute(sql, params).fetchall()
        for row in candidates:
            candidate = dict(row)
            job = get_job_requirement(candidate.get("role", ""), candidate.get("account_name", ""))
            if not job:
                skipped += 1
                continue
            result = match_candidate_with_requirement(candidate, job)
            try:
                next_raw = json.loads(result["raw_json"] or "{}")
            except Exception:
                next_raw = {}
            jd_dimension = ((next_raw.get("evaluation") or {}).get("dimensions") or {}).get("jd") or {}
            conn.execute(
                """
                UPDATE candidates
                SET raw_json = ?, score = ?, recommendation = ?, updated_at = ?
                WHERE id = ?
                """,
                (result["raw_json"], result["score"], result["recommendation"], now_iso(), candidate["id"]),
            )
            updated += 1
            if len(matched_items) < 100:
                matched_items.append({
                    "id": candidate.get("id"),
                    "name": candidate.get("name") or next_raw.get("name") or unresolved_candidate_display_name(candidate),
                    "role": candidate.get("role") or "",
                    "education": candidate.get("education") or next_raw.get("education") or "",
                    "experience": candidate.get("experience") or next_raw.get("experience") or "",
                    "expected_salary": candidate.get("expected_salary") or next_raw.get("expectedSalary") or "",
                    "account_name": candidate.get("account_name") or "",
                    "source": candidate.get("source") or "",
                    "score": result["score"],
                    "recommendation": result["recommendation"],
                    "jdMatch": jd_dimension.get("match") or "",
                    "matched": jd_dimension.get("matched") or [],
                    "missing": jd_dimension.get("missing") or [],
                })
    return {"success": True, "updated": updated, "skipped": skipped, "items": matched_items}


def save_recommendation(candidate: dict[str, Any], report: str = "") -> dict[str, Any]:
    candidate = normalize_candidate_payload(candidate)
    ok, reason = validate_candidate_minimum(candidate)
    if not ok:
        return {"success": False, "message": f"候选人推荐已阻止：{reason}"}
    if candidate.get("_invalidRecommendation"):
        return {"success": False, "message": candidate.get("_invalidReason", "候选人数据异常，已跳过推荐")}
    candidate_id = str(candidate.get("id") or candidate.get("candidate_id") or hashlib.sha1(json.dumps(candidate, ensure_ascii=False).encode()).hexdigest())
    rec_id = hashlib.sha1(f"{candidate_id}:{candidate.get('pushedAt') or now_iso()}".encode()).hexdigest()
    created_at = candidate.get("pushedAt") or now_iso()
    with connect() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO recommendations
            (id, candidate_id, name, role, source, account_name, score, recommendation, next_step, raw_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rec_id,
                candidate_id,
                candidate.get("name", ""),
                candidate.get("role", ""),
                candidate.get("source", ""),
                candidate.get("accountName", ""),
                int(candidate.get("score") or 0),
                candidate.get("recommendation", ""),
                candidate.get("nextStep", ""),
                json.dumps(candidate, ensure_ascii=False),
                created_at,
            ),
        )
        if report:
            report_id = hashlib.sha1(f"{candidate_id}:report:{report[:80]}".encode()).hexdigest()
            conn.execute(
                """
                INSERT OR REPLACE INTO reports
                (id, candidate_id, name, role, report, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (report_id, candidate_id, candidate.get("name", ""), candidate.get("role", ""), report, created_at),
            )
    return {"success": True, "id": rec_id}


def list_rows(
    table: str,
    limit: int = 200,
    date_field: str | None = None,
    date_value: str | None = None,
    filters: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    allowed = {"candidates", "recommendations", "reports", "job_requirements"}
    if table not in allowed:
        raise ValueError("invalid table")
    if table == "reports":
        sql = """
            SELECT
                r.*,
                COALESCE(NULLIF(c.source, ''), '未知来源') AS source,
                COALESCE(NULLIF(c.account_name, ''), '未识别') AS account_name
            FROM reports r
            LEFT JOIN candidates c ON c.id = r.candidate_id
        """
        order_field = "r.created_at"
    else:
        sql = f"SELECT * FROM {table}"
        order_field = "updated_at" if table == "job_requirements" else "created_at"
    params: list[Any] = []
    where: list[str] = []
    if date_field and date_value:
        qualified_date_field = f"r.{date_field}" if table == "reports" and "." not in date_field else date_field
        where.append(f"{qualified_date_field} LIKE ?")
        params.append(f"{date_value}%")
    filters = filters or {}
    q = (filters.get("q") or "").strip()
    source = (filters.get("source") or "").strip()
    account = (filters.get("account") or "").strip()
    account_exact = str(filters.get("accountExact") or filters.get("account_exact") or "").strip().lower() in {"1", "true", "yes"}
    if q:
        searchable = {
            "candidates": ["name", "role", "education", "experience", "expected_salary", "recommendation", "raw_json"],
            "recommendations": ["name", "role", "recommendation", "next_step", "raw_json"],
            "reports": ["r.name", "r.role", "r.report", "c.source", "c.account_name"],
            "job_requirements": ["role", "requirement", "source", "account_name"],
        }[table]
        where.append("(" + " OR ".join(f"{field} LIKE ?" for field in searchable) + ")")
        params.extend([f"%{q}%"] * len(searchable))
    if source:
        if table == "reports":
            where.append("COALESCE(NULLIF(c.source, ''), '未知来源') LIKE ?")
        elif table in {"candidates", "recommendations", "job_requirements"}:
            where.append("source LIKE ?")
        params.append(f"%{source}%")
    if account:
        operator = "=" if account_exact else "LIKE"
        if table == "reports":
            where.append(f"COALESCE(NULLIF(c.account_name, ''), '未识别') {operator} ?")
        elif table in {"candidates", "recommendations", "job_requirements"}:
            where.append(f"COALESCE(NULLIF(account_name, ''), '未识别') {operator} ?")
        params.append(account if account_exact else f"%{account}%")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += f" ORDER BY {order_field} DESC LIMIT ?"
    params.append(limit)
    with connect() as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def list_recommendation_details(
    limit: int = 1000,
    date_value: str | None = None,
    filters: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    where = []
    params: list[Any] = []
    if date_value:
        where.append("r.created_at LIKE ?")
        params.append(f"{date_value}%")
    filters = filters or {}
    q = str(filters.get("q") or "").strip()
    source = str(filters.get("source") or "").strip()
    account = str(filters.get("account") or "").strip()
    account_exact = str(filters.get("accountExact") or filters.get("account_exact") or "").strip().lower() in {"1", "true", "yes"}
    if q:
        fields = ["r.name", "r.role", "r.recommendation", "r.next_step", "r.raw_json", "c.education", "c.experience"]
        where.append("(" + " OR ".join(f"{field} LIKE ?" for field in fields) + ")")
        params.extend([f"%{q}%"] * len(fields))
    if source:
        where.append("COALESCE(NULLIF(c.source, ''), NULLIF(r.source, ''), '未知来源') LIKE ?")
        params.append(f"%{source}%")
    if account:
        operator = "=" if account_exact else "LIKE"
        where.append(f"COALESCE(NULLIF(c.account_name, ''), NULLIF(r.account_name, ''), '未识别') {operator} ?")
        params.append(account if account_exact else f"%{account}%")
    sql = """
        SELECT
            r.*,
            COALESCE(NULLIF(c.education, ''), '') AS education,
            COALESCE(NULLIF(c.experience, ''), '') AS experience,
            COALESCE(NULLIF(c.expected_salary, ''), '') AS expected_salary,
            COALESCE(NULLIF(c.source, ''), r.source) AS merged_source,
            COALESCE(NULLIF(c.account_name, ''), r.account_name) AS merged_account_name
        FROM recommendations r
        LEFT JOIN candidates c ON c.id = r.candidate_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY r.score DESC, r.created_at DESC LIMIT ?"
    params.append(limit)
    with connect() as conn:
        rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
    for row in rows:
        row["source"] = row.get("merged_source") or row.get("source") or ""
        row["account_name"] = row.get("merged_account_name") or row.get("account_name") or ""
    return rows


def parse_datetime_value(value: Any, end_of_day: bool = False) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) == 10 and re.match(r"\d{4}-\d{2}-\d{2}", text):
        suffix = "23:59:59" if end_of_day else "00:00:00"
        text = f"{text}T{suffix}"
    text = text.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(text)
        return dt.replace(tzinfo=None)
    except ValueError:
        return None


def row_datetime(row: dict[str, Any], *fields: str) -> datetime | None:
    for field in fields:
        dt = parse_datetime_value(row.get(field))
        if dt:
            return dt
    return None


def resolve_push_range(
    scope: str = "configured",
    start: str | None = None,
    end: str | None = None,
    settings: dict[str, Any] | None = None,
) -> tuple[datetime, datetime, str]:
    settings = settings or get_settings()
    mode = scope if scope and scope not in {"configured", "excel"} else str(settings.get("scheduledPushRangeMode") or "yesterday")
    now = datetime.now()
    if start or end:
        mode = "custom"
    if mode == "today":
        start_dt = datetime.combine(now.date(), datetime.min.time())
        end_dt = datetime.combine(now.date(), datetime.max.time()).replace(microsecond=0)
        label = "今日"
    elif mode == "all":
        start_dt = datetime(1970, 1, 1)
        end_dt = datetime.combine(now.date(), datetime.max.time()).replace(microsecond=0)
        label = "全部历史"
    elif mode == "last7":
        start_dt = datetime.combine((now - timedelta(days=6)).date(), datetime.min.time())
        end_dt = datetime.combine(now.date(), datetime.max.time()).replace(microsecond=0)
        label = "近7天"
    elif mode == "custom":
        start_dt = parse_datetime_value(start or settings.get("scheduledPushStart"), end_of_day=False)
        end_dt = parse_datetime_value(end or settings.get("scheduledPushEnd"), end_of_day=True)
        if not start_dt:
            start_dt = datetime.combine((now - timedelta(days=1)).date(), datetime.min.time())
        if not end_dt:
            end_dt = datetime.combine((now - timedelta(days=1)).date(), datetime.max.time()).replace(microsecond=0)
        if end_dt < start_dt:
            start_dt, end_dt = end_dt, start_dt
        label = f"{start_dt:%Y-%m-%d %H:%M} 至 {end_dt:%Y-%m-%d %H:%M}"
    else:
        day = now - timedelta(days=1)
        start_dt = datetime.combine(day.date(), datetime.min.time())
        end_dt = datetime.combine(day.date(), datetime.max.time()).replace(microsecond=0)
        label = "昨日"
    return start_dt, end_dt, label


def in_time_range(row: dict[str, Any], start_dt: datetime, end_dt: datetime, *fields: str) -> bool:
    dt = row_datetime(row, *fields)
    return bool(dt and start_dt <= dt <= end_dt)


def has_resume_evidence(row: dict[str, Any]) -> bool:
    candidate_id = str(row.get("candidate_id") or row.get("id") or "")
    raw = {}
    try:
        raw = json.loads(str(row.get("raw_json") or "{}"))
    except json.JSONDecodeError:
        raw = {}
    has_resume = bool(raw.get("hasResume"))
    if candidate_id.startswith("chat_") and not has_resume:
        return False
    if raw.get("hasResume") is False:
        return False
    if raw.get("resumeStatus") in {"未获取简历", "仅沟通信息"}:
        return False
    return bool(
        has_resume
        or raw.get("resumeStatus")
        or row.get("education")
        or row.get("experience")
        or raw.get("summary")
        or raw.get("rawText")
    )


def resume_request_satisfied(row: dict[str, Any]) -> bool:
    raw = raw_payload(row)
    if has_attachment_resume(row):
        return True
    score = int(row.get("score") or raw.get("score") or (raw.get("evaluation") or {}).get("score") or 0)
    if score < 40:
        return True
    status = str(raw.get("resumeRequestStatus") or "")
    return bool(raw.get("resumeRequestExecuted") or status in {"已点击求简历", "已索要简历", "已发送求简历消息"})


def qualifies_for_recommendation(row: dict[str, Any]) -> bool:
    raw = raw_payload(row)
    labels = {"推荐", "非常推荐", "强烈推荐"}
    score = int(row.get("score") or raw.get("score") or (raw.get("evaluation") or {}).get("score") or 0)
    recommendation = str(row.get("recommendation") or raw.get("recommendation") or (raw.get("evaluation") or {}).get("recommendation") or "")
    return score >= 40 or recommendation in labels


def get_range_dataset(
    scope: str = "configured",
    start: str | None = None,
    end: str | None = None,
    account: str | None = None,
) -> dict[str, Any]:
    settings = get_settings()
    start_dt, end_dt, label = resolve_push_range(scope, start, end, settings)
    filters = {"account": str(account or "").strip(), "accountExact": "1"} if str(account or "").strip() else {}
    candidates_all = list_rows("candidates", limit=100000, filters=filters)
    recommendations_all = list_recommendation_details(limit=100000, filters=filters)
    candidates = [
        item for item in candidates_all
        if in_time_range(item, start_dt, end_dt, "created_at", "updated_at")
    ]
    recommendations = [
        item for item in recommendations_all
        if in_time_range(item, start_dt, end_dt, "created_at")
    ]
    resume_candidates = [item for item in candidates if has_resume_evidence(item)]
    recommended_by_id: dict[str, dict[str, Any]] = {}
    for item in resume_candidates:
        if qualifies_for_recommendation(item):
            recommended_by_id[str(item.get("id") or "")] = item
    for item in [
        item for item in recommendations
        if has_resume_evidence(item) and qualifies_for_recommendation(item)
    ]:
        recommended_by_id[str(item.get("candidate_id") or item.get("id") or "")] = item
    recommended_candidates = list(recommended_by_id.values())
    return {
        "start": start_dt,
        "end": end_dt,
        "label": label,
        "account": str(account or "").strip(),
        "candidates": candidates,
        "resumeCandidates": resume_candidates,
        "recommendations": recommendations,
        "recommendedCandidates": recommended_candidates,
    }


def get_accounts() -> list[dict[str, Any]]:
    accounts: dict[str, dict[str, Any]] = {}

    def add_account(name_value: Any, count: int = 0, platform: str = "BOSS直聘", source: str = "data") -> None:
        name = account_label(name_value)
        key = normalize_account_name(name)
        if not key:
            return
        current = accounts.setdefault(key, {
            "name": name,
            "platform": platform or "BOSS直聘",
            "count": 0,
            "sources": set(),
            "manual": False,
        })
        if source == "data":
            current["count"] = int(current.get("count") or 0) + int(count or 0)
        elif not int(current.get("count") or 0):
            current["count"] = int(count or 0)
        current["platform"] = current.get("platform") or platform or "BOSS直聘"
        current["sources"].add(source)
        if source == "manual":
            current["manual"] = True

    with connect() as conn:
        candidate_rows = conn.execute(
            """
            SELECT
              COALESCE(NULLIF(account_name, ''), '未识别') AS name,
              COALESCE(MAX(NULLIF(account_platform, '')), 'BOSS直聘') AS platform,
              COUNT(*) AS count
            FROM candidates
            GROUP BY COALESCE(NULLIF(account_name, ''), '未识别')
            """
        ).fetchall()
        for row in candidate_rows:
            add_account(row["name"], int(row["count"] or 0), str(row["platform"] or "BOSS直聘"), "data")
        for table in ("recommendations", "job_requirements"):
            rows = conn.execute(
                f"""
                SELECT COALESCE(NULLIF(account_name, ''), '未识别') AS name, COUNT(*) AS count
                FROM {table}
                GROUP BY COALESCE(NULLIF(account_name, ''), '未识别')
                """
            ).fetchall()
            for row in rows:
                add_account(row["name"], int(row["count"] or 0), "BOSS直聘", table)
    settings = get_settings()
    for item in managed_accounts_from_settings(settings):
        add_account(item["name"], 0, item.get("platform", "BOSS直聘"), item.get("source", "manual"))
    return [
        {
            "name": item["name"],
            "platform": item.get("platform") or "BOSS直聘",
            "count": int(item.get("count") or 0),
            "manual": bool(item.get("manual")),
            "sources": sorted(item.get("sources") or []),
        }
        for item in sorted(accounts.values(), key=lambda item: (-int(item.get("count") or 0), item["name"]))
    ]


def get_stats(filters: dict[str, str] | None = None) -> dict[str, Any]:
    today = date_str()
    yesterday = date_str(-1)
    filters = filters or {}
    candidates = list_rows("candidates", limit=100000, filters=filters)
    recommendations = list_recommendation_details(limit=100000, filters=filters)
    reports = list_rows("reports", limit=100000, filters=filters)

    total_candidates = len(candidates)
    today_candidates = sum(1 for item in candidates if str(item.get("received_date") or "") == today)
    yesterday_candidates = sum(1 for item in candidates if str(item.get("received_date") or "") == yesterday)
    recommendation_count = len(recommendations)
    report_count = len(reports)
    scores = [int(item.get("score") or 0) for item in recommendations if str(item.get("score") or "").strip()]
    avg_score = (sum(scores) / len(scores)) if scores else 0

    source_counts: dict[str, int] = {}
    account_counts_map: dict[str, int] = {}
    for item in candidates:
        source = str(item.get("source") or "未知来源")
        account = account_label(item.get("account_name"))
        source_counts[source] = source_counts.get(source, 0) + 1
        account_counts_map[account] = account_counts_map.get(account, 0) + 1
    by_source = [
        {"name": name, "count": count}
        for name, count in sorted(source_counts.items(), key=lambda item: (-item[1], item[0]))
    ]
    by_account = [
        {"name": name, "count": count}
        for name, count in sorted(account_counts_map.items(), key=lambda item: (-item[1], item[0]))
    ]
    top_recommendations = sorted(
        recommendations,
        key=lambda item: (int(item.get("score") or 0), str(item.get("created_at") or "")),
        reverse=True,
    )[:10]
    return {
        "totalCandidates": total_candidates,
        "todayCandidates": today_candidates,
        "yesterdayCandidates": yesterday_candidates,
        "recommendationCount": recommendation_count,
        "reportCount": report_count,
        "averageScore": round(float(avg_score), 1),
        "bySource": by_source,
        "byAccount": by_account,
        "topRecommendations": top_recommendations,
    }


def rows_to_csv(rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([title for _, title in columns])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _ in columns])
    return output.getvalue()


def save_agent_conversation(
    question: str,
    answer: str,
    channel: str = "web",
    sender: str = "",
    raw: dict[str, Any] | None = None,
) -> dict[str, Any]:
    item_id = uuid.uuid4().hex
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO agent_conversations
            (id, channel, sender, question, answer, raw_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item_id,
                channel,
                sender,
                question,
                answer,
                json.dumps(raw or {}, ensure_ascii=False),
                now_iso(),
            ),
        )
    return {"success": True, "id": item_id}


def list_agent_conversations(limit: int = 100) -> list[dict[str, Any]]:
    with connect() as conn:
        return [
            dict(row) for row in conn.execute(
                "SELECT * FROM agent_conversations ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
        ]


def build_llm_history_context(question: str, max_items: int = 80, account: str | None = None) -> str:
    safe_items = max(10, min(int(max_items or 30), 80))
    filters = {"account": str(account or "").strip()} if str(account or "").strip() else {}
    candidates = list_rows("candidates", limit=safe_items, filters=filters)
    recommendations = list_recommendation_details(limit=safe_items, filters=filters)
    reports = list_rows("reports", limit=max(5, min(safe_items // 2, 20)), filters=filters)
    job_requirements = list_job_requirements(limit=max(10, min(safe_items, 40)), filters=filters)
    stats = get_stats(filters)
    parts = [
        "【统计概览】",
        f"账号范围：{account_label(account) if account else '全部账号'}",
        json.dumps(stats, ensure_ascii=False),
        "",
        "【候选人历史】",
    ]
    for idx, item in enumerate(candidates, 1):
        parts.append(
            f"{idx}. 姓名:{item.get('name','')}｜岗位:{item.get('role','')}｜学历:{item.get('education','')}｜"
            f"经验:{item.get('experience','')}｜薪资:{item.get('expected_salary','')}｜匹配度:{item.get('score',0)}｜"
            f"推荐:{item.get('recommendation','')}｜来源:{item.get('source','')}｜账号:{item.get('account_name','')}｜日期:{item.get('received_date','')}"
        )
    parts.extend(["", "【推荐记录】"])
    for idx, item in enumerate(recommendations, 1):
        parts.append(
            f"{idx}. 姓名:{item.get('name','')}｜岗位:{item.get('role','')}｜匹配度:{item.get('score',0)}｜"
            f"推荐:{item.get('recommendation','')}｜下一步:{item.get('next_step','')}｜来源:{item.get('source','')}｜账号:{item.get('account_name','')}｜时间:{item.get('created_at','')}"
        )
    parts.extend(["", "【候选人报告摘要】"])
    for idx, item in enumerate(reports[:30], 1):
        report = str(item.get("report") or "").replace("\n", " ")
        parts.append(f"{idx}. {item.get('name','')}｜{item.get('role','')}｜{report[:320]}")
    parts.extend(["", "【岗位要求库】"])
    for idx, item in enumerate(job_requirements[:80], 1):
        requirement = str(item.get("requirement") or "").replace("\n", " ")
        parts.append(f"{idx}. 岗位:{item.get('role','')}｜来源:{item.get('source','')}｜账号:{item.get('account_name','')}｜要求:{requirement[:360]}")
    return "\n".join(parts)[:max(3500, safe_items * 240)]


def build_llm_system_prompt() -> str:
    return (
        "你是招聘助手的问答 Agent。你只能根据提供的招聘历史数据回答问题。"
        "回答要使用中文，适合钉钉 markdown 展示。"
        "如果历史数据没有答案，要明确说没有找到，不要编造候选人。"
        "回答候选人时优先给出姓名、岗位、匹配度、推荐意见、来源、账号、下一步。"
    )


def request_json(
    url: str,
    payload: dict[str, Any],
    headers: dict[str, str],
    timeout: int = 30,
    retries: int = 0,
) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            **headers,
        },
        method="POST",
    )
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8") or "{}")
        except (TimeoutError, socket.timeout) as exc:
            last_error = exc
            if attempt >= retries:
                raise TimeoutError(f"请求读取超时（{timeout}秒）") from exc
            time.sleep(1.5 * (attempt + 1))
        except urllib.error.URLError as exc:
            last_error = exc
            if isinstance(exc.reason, (TimeoutError, socket.timeout)):
                if attempt >= retries:
                    raise TimeoutError(f"请求读取超时（{timeout}秒）") from exc
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
    raise RuntimeError(str(last_error or "请求失败"))


def request_llm_json(url: str, payload: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    timeout = max(30, min(int(config.get("llmTimeoutSeconds") or 90), 180))
    return request_json(
        url,
        payload,
        {"Authorization": f"Bearer {config['llmApiKey']}"},
        timeout=timeout,
        retries=1,
    )


def extract_anthropic_answer(body: dict[str, Any]) -> str:
    parts: list[str] = []
    for content in body.get("content", []) or []:
        if content.get("type") == "text" and content.get("text"):
            parts.append(str(content["text"]))
    return "\n".join(parts).strip()


def call_llm_chat_completions(question: str, context: str, config: dict[str, Any]) -> str:
    url = f"{str(config['llmApiBase']).rstrip('/')}/chat/completions"
    payload = {
        "model": config["llmModel"],
        "temperature": config["llmTemperature"],
        "max_tokens": int(config.get("llmMaxTokens") or 1000),
        "messages": [
            {"role": "system", "content": build_llm_system_prompt()},
            {"role": "user", "content": f"问题：{question}\n\n招聘历史数据：\n{context}"},
        ],
    }
    body = request_llm_json(url, payload, config)
    choices = body.get("choices") or []
    if not choices:
        raise RuntimeError("大模型 API 未返回 choices")
    message = choices[0].get("message") or {}
    answer = str(message.get("content") or "").strip()
    if not answer:
        raise RuntimeError("大模型 API 返回空答案")
    return answer[:18000]


def extract_responses_answer(body: dict[str, Any]) -> str:
    if body.get("output_text"):
        return str(body["output_text"]).strip()
    parts: list[str] = []
    for item in body.get("output", []) or []:
        for content in item.get("content", []) or []:
            if content.get("type") in {"output_text", "text"} and content.get("text"):
                parts.append(str(content["text"]))
    return "\n".join(parts).strip()


def call_llm_openai_responses(question: str, context: str, config: dict[str, Any]) -> str:
    url = f"{str(config['llmApiBase']).rstrip('/')}/responses"
    payload = {
        "model": config["llmModel"],
        "instructions": build_llm_system_prompt(),
        "input": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": f"问题：{question}\n\n招聘历史数据：\n{context}",
                    }
                ],
            }
        ],
        "max_output_tokens": int(config.get("llmMaxTokens") or 1000),
    }
    body = request_llm_json(url, payload, config)
    answer = extract_responses_answer(body)
    if not answer:
        raise RuntimeError("OpenAI Responses API 返回空答案")
    return answer[:18000]


def call_llm_anthropic_messages(question: str, context: str, config: dict[str, Any]) -> str:
    url = f"{str(config['llmApiBase']).rstrip('/')}/messages"
    payload = {
        "model": config["llmModel"],
        "max_tokens": int(config.get("llmMaxTokens") or 1000),
        "system": build_llm_system_prompt(),
        "messages": [
            {
                "role": "user",
                "content": f"问题：{question}\n\n招聘历史数据：\n{context}",
            }
        ],
    }
    body = request_json(
        url,
        payload,
        {
            "x-api-key": str(config["llmApiKey"]),
            "anthropic-version": "2023-06-01",
        },
        timeout=max(30, min(int(config.get("llmTimeoutSeconds") or 90), 180)),
        retries=1,
    )
    answer = extract_anthropic_answer(body)
    if not answer:
        raise RuntimeError("Claude Messages API 返回空答案")
    return answer[:18000]


def call_llm_chat(question: str, context: str) -> str:
    config = get_llm_config(mask_key=False)
    if not config.get("llmEnabled"):
        raise RuntimeError("大模型问答未启用")
    if not config.get("llmApiKey"):
        raise RuntimeError("大模型 API Key 未配置")
    if not config.get("llmApiBase"):
        raise RuntimeError("大模型 API Base URL 未配置")
    if not config.get("llmModel"):
        raise RuntimeError("大模型模型名未配置")

    try:
        if config.get("llmProtocol") == "openai-responses":
            return call_llm_openai_responses(question, context, config)
        if config.get("llmProtocol") == "anthropic-messages":
            return call_llm_anthropic_messages(question, context, config)
        return call_llm_chat_completions(question, context, config)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:1200]
        if exc.code == 401:
            provider_label = LLM_PROVIDER_PRESETS.get(str(config.get("llmProvider")), {}).get("label") or str(config.get("llmProvider") or "当前平台")
            key_url = LLM_PROVIDER_PRESETS.get(str(config.get("llmProvider")), {}).get("keyUrl") or "对应平台控制台"
            raise RuntimeError(
                "大模型 API 返回 401 Unauthorized。当前保存的 API Key 无效或不属于所选平台；"
                f"当前选择的是 {provider_label}，请在 {key_url} 获取并填写该平台 API Key。"
                f" 服务返回：{detail}"
            )
        if exc.code == 403:
            raise RuntimeError(
                "大模型 API 返回 403 Forbidden。请检查 API Key 是否有效、账户额度/模型权限是否开通、Base URL 是否正确。"
                f" 服务返回：{detail}"
            )
        raise RuntimeError(f"大模型 API HTTP {exc.code}: {detail}")
    except urllib.error.URLError as exc:
        raise RuntimeError(f"大模型 API 网络连接失败：{exc}")
    except TimeoutError as exc:
        raise RuntimeError(
            f"大模型 API 读取超时：{exc}。当前模型或平台响应较慢，请稍后重试，或在 Web 后台调低历史上下文条数/最大输出 tokens。"
        )
    except Exception:
        raise


PAGE_INTELLIGENCE_FIELDS = {
    "name", "role", "education", "experience", "expectedSalary", "ageGender",
    "currentCompany", "summary", "topLevelText", "jobRequirement", "resumeStatus",
    "confidence", "evidence",
}


def build_page_intelligence_system_prompt() -> str:
    return (
        "你是招聘助手的页面 OCR/NLP 识别引擎，负责从 BOSS 直聘页面文本和截图中抽取候选人信息。"
        "只允许根据输入材料抽取，不要编造。需要过滤导航栏、账号权益、职位管理、聊天历史、推荐列表等噪声。"
        "岗位 JD 只抽取“工作内容/工作职责/岗位职责/职位职责”和“工作要求/任职要求/岗位要求”下的介绍文字。"
        "必须只输出 JSON 对象，不要输出 markdown。"
    )


def build_page_intelligence_user_prompt(payload: dict[str, Any]) -> str:
    current = payload.get("currentCandidate") if isinstance(payload.get("currentCandidate"), dict) else {}
    text_blocks = payload.get("textBlocks") if isinstance(payload.get("textBlocks"), list) else []
    compact_blocks = []
    for item in text_blocks[:80]:
        if not isinstance(item, dict):
            continue
        compact_blocks.append({
            "text": str(item.get("text") or "")[:260],
            "x": item.get("x"),
            "y": item.get("y"),
            "w": item.get("w"),
            "h": item.get("h"),
        })
    context = {
        "task": str(payload.get("task") or "candidate-extract"),
        "url": str(payload.get("url") or ""),
        "currentCandidate": {
            key: current.get(key)
            for key in [
                "name", "role", "education", "experience", "expectedSalary", "ageGender",
                "currentCompany", "summary", "topLevelText", "jobRequirement", "resumeStatus",
            ]
            if current.get(key)
        },
        "candidatePanelText": str(payload.get("candidatePanelText") or "")[:3500],
        "topLevelText": str(payload.get("topLevelText") or "")[:3500],
        "roleContext": str(payload.get("roleContext") or "")[:2000],
        "pageText": str(payload.get("pageText") or "")[:12000],
        "textBlocks": compact_blocks,
    }
    schema = {
        "name": "候选人姓名，无法确认则空字符串",
        "role": "沟通职位/申请职位，保留 J 编号",
        "education": "学历，如 本科/硕士/大专",
        "experience": "工作经验，如 4年/应届",
        "expectedSalary": "期望薪资，如 20-30K",
        "ageGender": "年龄性别，如 28岁 男/女",
        "currentCompany": "当前或最近公司",
        "summary": "候选人顶层信息摘要，200字以内",
        "topLevelText": "只保留候选人详情顶层页面信息，不要混入下层页面/聊天历史",
        "jobRequirement": "如果页面展示 JD，仅提取工作内容和工作要求正文，否则空字符串",
        "resumeStatus": "有附件简历/无附件简历/仅沟通信息/未识别",
        "confidence": "0-100 的识别置信度",
        "evidence": ["用于判断的简短证据，最多5条"],
    }
    return (
        "请从以下页面材料中抽取候选人信息，并按指定 JSON schema 输出。\n"
        f"JSON schema 示例：{json.dumps(schema, ensure_ascii=False)}\n\n"
        f"页面材料：\n{json.dumps(context, ensure_ascii=False)}"
    )


def extract_json_object(text: str) -> dict[str, Any]:
    raw = str(text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        pass
    start = raw.find("{")
    if start < 0:
        return {}
    depth = 0
    in_string = False
    escape = False
    for index, char in enumerate(raw[start:], start):
        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                try:
                    parsed = json.loads(raw[start:index + 1])
                    return parsed if isinstance(parsed, dict) else {}
                except json.JSONDecodeError:
                    return {}
    return {}


def clean_page_intelligence_result(parsed: dict[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key in PAGE_INTELLIGENCE_FIELDS:
        value = parsed.get(key)
        if key == "confidence":
            try:
                cleaned[key] = max(0, min(100, int(float(value or 0))))
            except (TypeError, ValueError):
                cleaned[key] = 0
        elif key == "evidence":
            if isinstance(value, list):
                cleaned[key] = [str(item).strip()[:160] for item in value if str(item).strip()][:5]
            elif value:
                cleaned[key] = [str(value).strip()[:160]]
            else:
                cleaned[key] = []
        else:
            text = str(value or "").strip()
            if key in {"summary", "topLevelText"}:
                cleaned[key] = text[:1500]
            elif key == "jobRequirement":
                requirement = sanitize_job_requirement_text(text)
                cleaned[key] = requirement[:12000] if is_valid_job_requirement_text(requirement) else ""
            else:
                cleaned[key] = text[:260]
    return cleaned


def extract_page_intelligence(payload: dict[str, Any]) -> dict[str, Any]:
    config = get_llm_config(mask_key=False)
    if not config.get("llmEnabled") or not config.get("llmApiKey"):
        return {
            "success": True,
            "skipped": True,
            "message": "大模型未启用或 API Key 未配置，已跳过页面智能识别",
            "extracted": {},
        }
    if config.get("llmProtocol") not in {"openai-chat", "openai-responses"}:
        return {
            "success": True,
            "skipped": True,
            "message": "页面智能识别目前使用 OpenAI-compatible/Responses 模式，请选择智谱 GLM、硅基流动或 OpenAI-compatible 平台",
            "extracted": {},
        }

    screenshot = str(payload.get("screenshotDataUrl") or "").strip()
    if screenshot and (len(screenshot) > 4 * 1024 * 1024 or not screenshot.startswith("data:image/")):
        screenshot = ""
    system_prompt = build_page_intelligence_system_prompt()
    user_prompt = build_page_intelligence_user_prompt(payload)

    if config.get("llmProtocol") == "openai-responses":
        content: list[dict[str, Any]] = [{"type": "input_text", "text": user_prompt}]
        if screenshot:
            content.append({"type": "input_image", "image_url": screenshot})
        request_payload = {
            "model": config["llmModel"],
            "instructions": system_prompt,
            "input": [{"role": "user", "content": content}],
            "max_output_tokens": min(int(config.get("llmMaxTokens") or 1000), 1800),
        }
        body = request_llm_json(f"{str(config['llmApiBase']).rstrip('/')}/responses", request_payload, config)
        answer = extract_responses_answer(body)
    else:
        content: Any = user_prompt
        if screenshot:
            content = [
                {"type": "text", "text": user_prompt},
                {"type": "image_url", "image_url": {"url": screenshot}},
            ]
        request_payload = {
            "model": config["llmModel"],
            "temperature": min(float(config.get("llmTemperature") or 0.1), 0.3),
            "max_tokens": min(int(config.get("llmMaxTokens") or 1000), 1800),
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": content},
            ],
        }
        body = request_llm_json(f"{str(config['llmApiBase']).rstrip('/')}/chat/completions", request_payload, config)
        choices = body.get("choices") or []
        if not choices:
            raise RuntimeError("页面智能识别模型未返回 choices")
        answer = str((choices[0].get("message") or {}).get("content") or "").strip()

    parsed = extract_json_object(answer)
    if not parsed:
        raise RuntimeError("页面智能识别模型未返回有效 JSON")
    return {
        "success": True,
        "skipped": False,
        "provider": config.get("llmProvider"),
        "model": config.get("llmModel"),
        "usedScreenshot": bool(screenshot),
        "extracted": clean_page_intelligence_result(parsed),
    }


def should_use_fast_rules(question: str) -> bool:
    normalized = str(question or "").strip()
    if not normalized:
        return True
    fast_keywords = [
        "多少", "数量", "统计", "汇总", "昨天", "昨日", "今天", "今日",
        "推荐候选人", "收到简历", "新增候选人", "数据来源", "账号分布",
        "有哪些", "是谁", "最高", "排名", "最合适",
    ]
    analysis_keywords = ["分析原因", "为什么", "怎么优化", "建议", "写一段", "生成话术", "对比分析"]
    return any(keyword in normalized for keyword in fast_keywords) and not any(
        keyword in normalized for keyword in analysis_keywords
    )


def answer_question_with_agent(question: str, account: str | None = None, force_llm: bool = False) -> tuple[str, dict[str, Any]]:
    fallback = answer_question_rules(question, account)
    config = get_llm_config(mask_key=False)
    provider = str(config.get("llmProvider") or DEFAULT_LLM_CONFIG["llmProvider"])
    provider_label = LLM_PROVIDER_PRESETS.get(provider, {}).get("label") or provider
    if not config.get("llmEnabled"):
        return fallback, {"mode": "rules", "llmEnabled": False, "provider": provider, "providerLabel": provider_label}
    try:
        context = build_llm_history_context(question, int(config.get("llmMaxContextItems") or 80), account)
        answer = call_llm_chat(question, context)
        return answer, {
            "mode": "llm",
            "provider": provider,
            "providerLabel": provider_label,
            "model": config.get("llmModel"),
            "contextLength": len(context),
        }
    except Exception as exc:
        return (
            f"{fallback}\n\n---\n\n> 大模型回答暂不可用，已使用本地历史规则回答。原因：{exc}",
            {
                "mode": "rules-fallback",
                "error": str(exc),
                "provider": provider,
                "providerLabel": provider_label,
                "model": config.get("llmModel"),
            },
        )


def get_settings() -> dict[str, Any]:
    with connect() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    result: dict[str, Any] = {}
    for row in rows:
        try:
            result[row["key"]] = json.loads(row["value"])
        except json.JSONDecodeError:
            result[row["key"]] = row["value"]
    return result


def save_settings(payload: dict[str, Any]) -> dict[str, Any]:
    current = get_settings()
    account_name = str(payload.get("accountName") or "").strip()
    if account_name:
        item = normalize_managed_account({
            "name": account_name,
            "platform": payload.get("accountPlatform") or current.get("accountPlatform") or "BOSS直聘",
            "source": "manual" if payload.get("accountNameManual") else "current",
        })
        if item:
            accounts = managed_accounts_from_settings({**current, **payload})
            by_key = {normalize_account_name(account["name"]): account for account in accounts}
            by_key[normalize_account_name(item["name"])] = item
            payload = {**payload, "managedAccounts": sorted(by_key.values(), key=lambda account: account["name"])}
    with connect() as conn:
        for key, value in payload.items():
            if key in {"dingtalkAppKey", "dingtalkAppSecret"} and str(value or "").strip() in {"", "********"} and current.get(key):
                value = current.get(key)
            conn.execute(
                "INSERT OR REPLACE INTO settings (key, value, updated_at) VALUES (?, ?, ?)",
                (key, json.dumps(value, ensure_ascii=False), now_iso()),
            )
    return {"success": True}


def get_llm_config(mask_key: bool = False) -> dict[str, Any]:
    settings = get_settings()
    provider = str(settings.get("llmProvider") or DEFAULT_LLM_CONFIG["llmProvider"])
    preset = LLM_PROVIDER_PRESETS.get(provider, LLM_PROVIDER_PRESETS["custom"])
    config = {
        **DEFAULT_LLM_CONFIG,
        "llmEnabled": bool(settings.get("llmEnabled", DEFAULT_LLM_CONFIG["llmEnabled"])),
        "llmProvider": provider,
        "llmProtocol": str(settings.get("llmProtocol") or preset["protocol"]),
        "llmApiBase": str(settings.get("llmApiBase") or preset.get("apiBase") or "").rstrip("/"),
        "llmApiKey": str(settings.get("llmApiKey") or ""),
        "llmModel": str(settings.get("llmModel") or preset.get("model") or ""),
        "llmTemperature": float(settings.get("llmTemperature", DEFAULT_LLM_CONFIG["llmTemperature"]) or 0.2),
        "llmMaxContextItems": int(settings.get("llmMaxContextItems", DEFAULT_LLM_CONFIG["llmMaxContextItems"]) or 80),
        "llmMaxTokens": int(settings.get("llmMaxTokens", DEFAULT_LLM_CONFIG["llmMaxTokens"]) or 1000),
        "llmTimeoutSeconds": int(settings.get("llmTimeoutSeconds", DEFAULT_LLM_CONFIG["llmTimeoutSeconds"]) or 90),
    }
    if mask_key and config.get("llmApiKey"):
        config["llmApiKey"] = "********"
        config["llmApiKeyConfigured"] = True
    else:
        config["llmApiKeyConfigured"] = bool(config.get("llmApiKey"))
    return config


def save_llm_config(payload: dict[str, Any]) -> dict[str, Any]:
    current = get_llm_config(mask_key=False)
    provider = str(payload.get("llmProvider") or current["llmProvider"] or DEFAULT_LLM_CONFIG["llmProvider"])
    preset = LLM_PROVIDER_PRESETS.get(provider, LLM_PROVIDER_PRESETS["custom"])
    api_base = str(payload.get("llmApiBase") or preset.get("apiBase") or "").rstrip("/")
    model = str(payload.get("llmModel") or preset.get("model") or "")
    config: dict[str, Any] = {
        "llmEnabled": bool(payload.get("llmEnabled", current["llmEnabled"])),
        "llmProvider": provider,
        "llmProtocol": preset["protocol"],
        "llmApiBase": api_base,
        "llmModel": model,
        "llmTemperature": max(0, min(float(payload.get("llmTemperature", current["llmTemperature"]) or 0.2), 2)),
        "llmMaxContextItems": max(10, min(int(payload.get("llmMaxContextItems", current["llmMaxContextItems"]) or 80), 500)),
        "llmMaxTokens": max(100, min(int(payload.get("llmMaxTokens", current["llmMaxTokens"]) or 1000), 8000)),
        "llmTimeoutSeconds": max(30, min(int(payload.get("llmTimeoutSeconds", current.get("llmTimeoutSeconds", 90)) or 90), 180)),
    }
    if "pageIntelligenceEnabled" in payload:
        config["pageIntelligenceEnabled"] = bool(payload.get("pageIntelligenceEnabled"))
    if "pageIntelligenceUseScreenshot" in payload:
        config["pageIntelligenceUseScreenshot"] = bool(payload.get("pageIntelligenceUseScreenshot"))
    api_key = str(payload.get("llmApiKey") or "").strip()
    if api_key and api_key != "********":
        config["llmApiKey"] = api_key
    elif current.get("llmApiKey"):
        config["llmApiKey"] = current["llmApiKey"]
    else:
        config["llmApiKey"] = ""
    save_settings(config)
    return {"success": True, "llm": get_llm_config(mask_key=True)}


def reset_llm_config() -> dict[str, Any]:
    save_settings({
        "llmEnabled": DEFAULT_LLM_CONFIG["llmEnabled"],
        "llmProvider": DEFAULT_LLM_CONFIG["llmProvider"],
        "llmProtocol": DEFAULT_LLM_CONFIG["llmProtocol"],
        "llmApiBase": DEFAULT_LLM_CONFIG["llmApiBase"],
        "llmApiKey": "",
        "llmModel": DEFAULT_LLM_CONFIG["llmModel"],
        "llmTemperature": DEFAULT_LLM_CONFIG["llmTemperature"],
        "llmMaxContextItems": DEFAULT_LLM_CONFIG["llmMaxContextItems"],
        "llmMaxTokens": DEFAULT_LLM_CONFIG["llmMaxTokens"],
        "llmTimeoutSeconds": DEFAULT_LLM_CONFIG["llmTimeoutSeconds"],
        "pageIntelligenceEnabled": True,
        "pageIntelligenceUseScreenshot": False,
    })
    return {"success": True, "llm": get_llm_config(mask_key=True)}


def behavior_policy_key(account: Any) -> str:
    return normalize_account_name(account_label(account))


def get_behavior_policy(account: str = "") -> dict[str, Any]:
    settings = get_settings()
    saved = settings.get("behaviorPolicy") if isinstance(settings.get("behaviorPolicy"), dict) else {}
    account_name = account_label(account) if str(account or "").strip() else ""
    account_map = settings.get("behaviorPoliciesByAccount") if isinstance(settings.get("behaviorPoliciesByAccount"), dict) else {}
    account_saved = account_map.get(behavior_policy_key(account_name), {}) if account_name else {}
    if not isinstance(account_saved, dict):
        account_saved = {}
    policy = {**DEFAULT_BEHAVIOR_POLICY, **saved, **account_saved}
    interaction_modes = {
        **DEFAULT_BEHAVIOR_POLICY["interactionModes"],
        **(saved.get("interactionModes", {}) if isinstance(saved.get("interactionModes"), dict) else {}),
        **(account_saved.get("interactionModes", {}) if isinstance(account_saved.get("interactionModes"), dict) else {}),
    }
    policy["interactionModes"] = interaction_modes
    policy["pageIntelligenceEnabled"] = bool(settings.get("pageIntelligenceEnabled", True))
    policy["pageIntelligenceUseScreenshot"] = bool(settings.get("pageIntelligenceUseScreenshot", False))
    policy["accountName"] = account_name
    policy["scope"] = "account" if account_name else "global"
    return policy


def save_behavior_policy(payload: dict[str, Any]) -> dict[str, Any]:
    account_name = str(payload.get("accountName") or payload.get("account") or "").strip()
    current = get_behavior_policy(account_name)
    next_policy = {**current}
    numeric_fields = [
        "requestDelayMin", "requestDelayMax", "detailDwellMin", "detailDwellMax",
        "actionDwellMin", "actionDwellMax", "dailyLimit", "hourlyLimit",
        "maxCandidatesPerRun", "browseProbability", "longBreakEvery",
        "longBreakMin", "longBreakMax",
    ]
    for field in numeric_fields:
        if field in payload:
            value = float(payload[field]) if field == "browseProbability" else int(payload[field])
            next_policy[field] = max(0, value)
    if next_policy["requestDelayMax"] < next_policy["requestDelayMin"]:
        next_policy["requestDelayMax"] = next_policy["requestDelayMin"]
    if next_policy["detailDwellMax"] < next_policy["detailDwellMin"]:
        next_policy["detailDwellMax"] = next_policy["detailDwellMin"]
    if next_policy["actionDwellMax"] < next_policy["actionDwellMin"]:
        next_policy["actionDwellMax"] = next_policy["actionDwellMin"]

    next_policy["behaviorPolicyEnabled"] = bool(payload.get("behaviorPolicyEnabled", next_policy.get("behaviorPolicyEnabled")))
    if payload.get("scrollMode") in {"mixed", "fast", "slow", "segmented"}:
        next_policy["scrollMode"] = payload["scrollMode"]
    if isinstance(payload.get("interactionModes"), dict):
        next_policy["interactionModes"] = {
            "manualPage": int(payload["interactionModes"].get("manualPage", 40) or 0),
            "detailClick": int(payload["interactionModes"].get("detailClick", 35) or 0),
            "filterReview": int(payload["interactionModes"].get("filterReview", 25) or 0),
        }
    if isinstance(payload.get("searchKeywordPool"), list):
        next_policy["searchKeywordPool"] = [
            str(item).strip() for item in payload["searchKeywordPool"] if str(item).strip()
        ][:30]
    elif isinstance(payload.get("searchKeywordPool"), str):
        next_policy["searchKeywordPool"] = [
            item.strip() for item in payload["searchKeywordPool"].replace("，", "\n").splitlines() if item.strip()
        ][:30]
    next_policy["workTimeEnabled"] = bool(payload.get("workTimeEnabled", next_policy.get("workTimeEnabled")))
    for field in ("workStartTime", "workEndTime"):
        value = str(payload.get(field) or next_policy.get(field) or "").strip()
        if len(value) >= 5 and value[2] == ":":
            next_policy[field] = value[:5]
    if isinstance(payload.get("workDays"), list):
        next_policy["workDays"] = [
            day for day in [int(item) for item in payload["workDays"] if str(item).isdigit()]
            if 1 <= day <= 7
        ] or DEFAULT_BEHAVIOR_POLICY["workDays"]

    next_policy.pop("scope", None)
    next_policy["accountName"] = account_name
    if account_name:
        settings = get_settings()
        account_map = settings.get("behaviorPoliciesByAccount") if isinstance(settings.get("behaviorPoliciesByAccount"), dict) else {}
        account_map[behavior_policy_key(account_name)] = next_policy
        save_settings({"behaviorPoliciesByAccount": account_map})
    else:
        next_policy.pop("accountName", None)
        save_settings({"behaviorPolicy": next_policy})
    return {"success": True, "behaviorPolicy": next_policy}


def scheduled_push_loop() -> None:
    while True:
        try:
            settings = get_settings()
            enabled = bool(settings.get("scheduledPushEnabled"))
            push_time = str(settings.get("scheduledPushTime") or "10:00")
            last_date = str(settings.get("scheduledPushLastDate") or "")
            today = date_str()
            if enabled and last_date != today and datetime.now().strftime("%H:%M") == push_time:
                scope = str(settings.get("scheduledPushRangeMode") or "yesterday")
                output_path, dataset = create_recommendation_excel(scope)
                markdown = build_push_markdown(scope, dataset, output_path)
                result = send_dingtalk_markdown("招聘助手候选人简历汇总", markdown)
                delivery_result = deliver_dingtalk_excel(output_path, dataset)
                if result.get("success") and delivery_result.get("success"):
                    save_settings({"scheduledPushLastDate": today})
                elif not delivery_result.get("success"):
                    print(f"[scheduled_push] Excel 推送失败：{delivery_result.get('message') or delivery_result}")
        except Exception as exc:
            print(f"[scheduled_push] {exc}")
        time.sleep(60)


def request_base_url(handler: SimpleHTTPRequestHandler) -> str:
    forwarded_proto = handler.headers.get("X-Forwarded-Proto")
    forwarded_host = handler.headers.get("X-Forwarded-Host")
    proto = forwarded_proto or ("https" if handler.headers.get("X-Forwarded-Ssl") == "on" else "http")
    host = forwarded_host or handler.headers.get("Host", "")
    return f"{proto}://{host}".rstrip("/") if host else ""


def usable_base_url(value: Any) -> str:
    base = str(value or "").strip().rstrip("/")
    if not base:
        return ""
    parsed = urllib.parse.urlparse(base)
    host = (parsed.hostname or "").lower()
    if host in {"hr.example.com", "example.com"} or host.endswith(".example.com"):
        return ""
    return base if parsed.scheme in {"http", "https"} and parsed.netloc else ""


def configured_backend_url(handler: SimpleHTTPRequestHandler, override: str = "") -> str:
    candidates = [
        override,
        request_base_url(handler),
        usable_base_url(get_settings().get("adminBaseUrl")),
        "http://10.100.60.5:8787",
    ]
    for value in candidates:
        base = usable_base_url(value)
        if base:
            return base
    raise ValueError("无法识别插件后端地址")


def extension_match_pattern(base_url: str) -> str:
    parsed = urllib.parse.urlparse(base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("插件后端地址必须是 http/https URL")
    host = parsed.hostname
    if ":" in host and not host.startswith("["):
        host = f"[{host}]"
    return f"{parsed.scheme}://{host}/*"


def normalize_extension_permission(pattern: str) -> str:
    return (
        str(pattern or "")
        .replace("http://localhost:*/*", "http://localhost/*")
        .replace("http://127.0.0.1:*/*", "http://127.0.0.1/*")
    )


def extension_package_filename(base_url: str) -> str:
    parsed = urllib.parse.urlparse(base_url)
    host = (parsed.hostname or "backend").replace(":", "_")
    port = f"_{parsed.port}" if parsed.port else ""
    return f"招聘助手插件_{host}{port}.zip"


def patch_extension_javascript(text: str, backend_url: str) -> str:
    return re.sub(
        r"backendUrl:\s*'[^']*'",
        f"backendUrl: '{backend_url}'",
        text,
    )


def build_configured_extension_package(base_url: str) -> tuple[bytes, dict[str, Any]]:
    if not EXTENSION_DIR.exists():
        raise FileNotFoundError(f"浏览器插件目录不存在：{EXTENSION_DIR}")

    backend_url = usable_base_url(base_url)
    if not backend_url:
        raise ValueError("插件后端地址无效")
    backend_permission = extension_match_pattern(backend_url)
    manifest_path = EXTENSION_DIR / "manifest.json"
    manifest = json.loads(manifest_path.read_text("utf-8"))
    permissions = list(dict.fromkeys(
        item for item in [*(normalize_extension_permission(p) for p in manifest.get("host_permissions", [])), backend_permission]
        if item
    ))
    manifest["host_permissions"] = permissions

    install_notes = f"""招聘助手浏览器插件安装说明

后端服务地址：{backend_url}

安装步骤：
1. 解压本 zip 文件。
2. Chrome 或 Edge 打开扩展管理页面。
3. 开启开发者模式。
4. 选择“加载已解压的扩展程序”，选择解压后的文件夹。
5. 打开 BOSS 直聘沟通页面，点击插件“一键开始执行任务”。

本安装包已内置后端地址和浏览器访问权限，安装后无需再填写后端配置。
"""

    config = {
        "backendUrl": backend_url,
        "generatedAt": now_iso(),
        "hostPermission": backend_permission,
        "source": "web-admin",
    }
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(EXTENSION_DIR.rglob("*")):
            if path.is_dir() or path.name == ".DS_Store":
                continue
            rel = path.relative_to(EXTENSION_DIR).as_posix()
            if rel == "manifest.json":
                archive.writestr(rel, json.dumps(manifest, ensure_ascii=False, indent=2))
                continue
            if rel in {"background.js", "popup.js"}:
                archive.writestr(rel, patch_extension_javascript(path.read_text("utf-8"), backend_url))
                continue
            archive.write(path, rel)
        archive.writestr("extension-config.json", json.dumps(config, ensure_ascii=False, indent=2))
        archive.writestr("安装说明.txt", install_notes)

    return buffer.getvalue(), {
        "backendUrl": backend_url,
        "filename": extension_package_filename(backend_url),
        "hostPermission": backend_permission,
        "generatedAt": config["generatedAt"],
    }


def export_url(path: Path, base_url: str | None = None) -> str:
    settings = get_settings()
    base = (
        usable_base_url(settings.get("publicBaseUrl"))
        or usable_base_url(os.environ.get("PUBLIC_BASE_URL"))
        or usable_base_url(base_url)
    )
    if not base:
        base = (
            usable_base_url(DEFAULT_PUBLIC_BASE_URL)
            or usable_base_url(settings.get("adminBaseUrl"))
        )
    filename = urllib.parse.quote(path.name)
    return f"{base}/exports/{filename}" if base else f"/exports/{filename}"


def is_public_http_url(value: str) -> bool:
    parsed = urllib.parse.urlparse(str(value or ""))
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def build_push_markdown(scope: str, dataset: dict[str, Any], excel_path: Path, base_url: str | None = None) -> str:
    summary = build_summary(
        "custom",
        dataset["start"].isoformat(timespec="minutes"),
        dataset["end"].isoformat(timespec="minutes"),
        dataset.get("account") or "",
    )
    excel_link = export_url(excel_path, base_url)
    file_line = (
        f"- 文件：[{excel_path.name}]({excel_link})"
        if is_public_http_url(excel_link)
        else f"- 文件：{excel_path.name}"
    )
    excel_section = "\n".join([
        "#### Excel 推荐表",
        file_line,
        f"- 范围：{dataset['label']}（{dataset['start']:%Y-%m-%d %H:%M} 至 {dataset['end']:%Y-%m-%d %H:%M}）",
        "- 口径：仅推荐已识别到简历证据且匹配度达标的候选人；未获取简历的沟通记录已排除。",
    ])
    return f"{summary}\n\n{excel_section}"


def dingtalk_signed_url(webhook: str, secret: str = "") -> str:
    if not secret:
        return webhook
    timestamp = str(int(time.time() * 1000))
    string_to_sign = f"{timestamp}\n{secret}".encode("utf-8")
    digest = hmac.new(secret.encode("utf-8"), string_to_sign, hashlib.sha256).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(digest).decode("utf-8"))
    separator = "&" if "?" in webhook else "?"
    return f"{webhook}{separator}timestamp={timestamp}&sign={sign}"


def dingtalk_ssl_context() -> ssl.SSLContext:
    context = ssl.create_default_context()
    if hasattr(ssl, "TLSVersion"):
        context.minimum_version = ssl.TLSVersion.TLSv1_2
    return context


def is_transient_dingtalk_network_error(exc: BaseException) -> bool:
    if isinstance(exc, (TimeoutError, socket.timeout, ssl.SSLError, ConnectionResetError, http.client.RemoteDisconnected)):
        return True
    if isinstance(exc, urllib.error.URLError):
        return is_transient_dingtalk_network_error(exc.reason)
    text = str(exc)
    return any(marker in text for marker in (
        "UNEXPECTED_EOF_WHILE_READING",
        "EOF occurred in violation of protocol",
        "Connection reset",
        "Remote end closed connection",
        "read operation timed out",
    ))


def dingtalk_request_json(
    url: str,
    payload: dict[str, Any] | None = None,
    headers: dict[str, str] | None = None,
    timeout: int = 15,
    retries: int = 2,
    method: str | None = None,
) -> dict[str, Any]:
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request_method = method or ("POST" if data is not None else "GET")
    base_headers = {
        "User-Agent": "RecruitmentBot/1.0",
        "Accept": "application/json",
        "Connection": "close",
        **(headers or {}),
    }
    if data is not None and "Content-Type" not in base_headers:
        base_headers["Content-Type"] = "application/json;charset=utf-8"

    last_error: BaseException | None = None
    for attempt in range(retries + 1):
        request = urllib.request.Request(url, data=data, headers=base_headers, method=request_method)
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=dingtalk_ssl_context()) as response:
                return json.loads(response.read().decode("utf-8") or "{}")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="ignore")
            try:
                parsed = json.loads(body or "{}")
            except json.JSONDecodeError:
                parsed = {"message": body or str(exc)}
            parsed.setdefault("httpStatus", exc.code)
            return parsed
        except Exception as exc:
            last_error = exc
            if attempt >= retries or not is_transient_dingtalk_network_error(exc):
                break
            time.sleep(min(6, 1.2 * (attempt + 1)))

    raise RuntimeError(
        "钉钉 HTTPS 连接被部署环境中断："
        f"{last_error}. 请检查服务器到 oapi.dingtalk.com/api.dingtalk.com 的 443 出站访问、HTTPS 代理和 CA 证书。"
    )


def dingtalk_request_bytes(
    url: str,
    data: bytes,
    headers: dict[str, str],
    timeout: int = 30,
    retries: int = 2,
) -> dict[str, Any]:
    base_headers = {
        "User-Agent": "RecruitmentBot/1.0",
        "Accept": "application/json",
        "Connection": "close",
        **headers,
    }
    last_error: BaseException | None = None
    for attempt in range(retries + 1):
        request = urllib.request.Request(url, data=data, headers=base_headers, method="POST")
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=dingtalk_ssl_context()) as response:
                return json.loads(response.read().decode("utf-8") or "{}")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="ignore")
            try:
                parsed = json.loads(body or "{}")
            except json.JSONDecodeError:
                parsed = {"message": body or str(exc)}
            parsed.setdefault("httpStatus", exc.code)
            return parsed
        except Exception as exc:
            last_error = exc
            if attempt >= retries or not is_transient_dingtalk_network_error(exc):
                break
            time.sleep(min(6, 1.2 * (attempt + 1)))
    raise RuntimeError(
        "钉钉 HTTPS 文件请求被部署环境中断："
        f"{last_error}. 请检查服务器出站网络、HTTPS 代理和 CA 证书。"
    )


def dingtalk_code_ok(value: Any) -> bool:
    return value in (None, "", 0, "0")


def dingtalk_error_message(body: Any, default: str = "钉钉接口调用失败") -> str:
    if not isinstance(body, dict):
        return default
    for key in ("errmsg", "message", "msg", "errorMessage", "error"):
        value = body.get(key)
        if value and str(value).lower() not in {"ok", "success"}:
            return str(value)
    return default


def dingtalk_response_success(body: Any) -> bool:
    if not isinstance(body, dict):
        return False
    try:
        if int(body.get("httpStatus") or 0) >= 400:
            return False
    except (TypeError, ValueError):
        return False
    if not dingtalk_code_ok(body.get("errcode")) or not dingtalk_code_ok(body.get("code")):
        return False
    for key in ("errmsg", "message"):
        value = body.get(key)
        if value and str(value).lower() not in {"ok", "success"}:
            return False
    return True


def send_dingtalk_markdown(title: str, text: str) -> dict[str, Any]:
    settings = get_settings()
    webhook = str(settings.get("dingtalkWebhook") or "")
    secret = str(settings.get("dingtalkSecret") or "")
    if not webhook:
        return {"success": False, "message": "钉钉 Webhook 未配置"}
    return send_dingtalk_markdown_to_webhook(webhook, title, text, secret)


def send_dingtalk_markdown_to_webhook(webhook: str, title: str, text: str, secret: str = "") -> dict[str, Any]:
    url = dingtalk_signed_url(webhook, secret)
    try:
        body = dingtalk_request_json(url, {"msgtype": "markdown", "markdown": {"title": title, "text": text}}, timeout=15, retries=2)
    except RuntimeError as exc:
        return {"success": False, "message": str(exc)}
    if not dingtalk_response_success(body):
        return {"success": False, "message": dingtalk_error_message(body, "钉钉推送失败"), "body": body}
    return {"success": True, "body": body}


def get_dingtalk_access_token(settings: dict[str, Any]) -> str:
    app_key = str(settings.get("dingtalkAppKey") or "").strip()
    app_secret = str(settings.get("dingtalkAppSecret") or "").strip()
    if not app_key or not app_secret:
        return ""
    url = "https://oapi.dingtalk.com/gettoken?" + urllib.parse.urlencode({
        "appkey": app_key,
        "appsecret": app_secret,
    })
    body = dingtalk_request_json(url, timeout=20, retries=2, method="GET")
    if not dingtalk_response_success(body):
        raise RuntimeError(f"获取钉钉 access_token 失败：{dingtalk_error_message(body, str(body))}")
    token = str(body.get("access_token") or "")
    if not token:
        raise RuntimeError(f"获取钉钉 access_token 失败：接口未返回 access_token，返回内容：{body}")
    return token


def multipart_form_data(fields: dict[str, str], files: dict[str, tuple[str, bytes, str]]) -> tuple[bytes, str]:
    boundary = f"----RecruitmentBot{uuid.uuid4().hex}"
    chunks: list[bytes] = []
    for name, value in fields.items():
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
            str(value).encode("utf-8"),
            b"\r\n",
        ])
    for name, (filename, data, content_type) in files.items():
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode(),
            f"Content-Type: {content_type}\r\n\r\n".encode(),
            data,
            b"\r\n",
        ])
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def dingtalk_upload_filename(file_path: Path) -> str:
    suffix = file_path.suffix if file_path.suffix.lower() == ".xlsx" else ".xlsx"
    timestamp = re.search(r"\d{8}_\d{6}", file_path.name)
    if timestamp:
        return f"recommendation_{timestamp.group(0)}{suffix}"
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", file_path.stem).strip("._-")
    return f"{(stem or 'recommendation')[:80]}{suffix}"


def upload_dingtalk_file(access_token: str, file_path: Path) -> str:
    upload_name = dingtalk_upload_filename(file_path)
    data, content_type = multipart_form_data(
        {},
        {
            "media": (
                upload_name,
                file_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    url = f"https://oapi.dingtalk.com/media/upload?access_token={urllib.parse.quote(access_token)}&type=file"
    body = dingtalk_request_bytes(
        url,
        data,
        {"Content-Type": content_type},
        timeout=45,
        retries=2,
    )
    media_id = str(body.get("media_id") or body.get("mediaId") or "")
    if not dingtalk_response_success(body) or not media_id:
        raise RuntimeError(f"上传钉钉文件失败：{dingtalk_error_message(body, str(body))}")
    return media_id


def send_dingtalk_file(file_path: Path) -> dict[str, Any]:
    settings = get_settings()
    try:
        access_token = get_dingtalk_access_token(settings)
        if not access_token:
            return {"success": False, "skipped": True, "message": "未配置钉钉 AppKey/AppSecret，无法直接发送文件"}
        media_id = upload_dingtalk_file(access_token, file_path)
    except RuntimeError as exc:
        return {"success": False, "message": str(exc)}
    open_conversation_id = str(settings.get("dingtalkOpenConversationId") or "").strip()
    robot_code = str(settings.get("dingtalkRobotCode") or settings.get("dingtalkAppKey") or "").strip()
    if open_conversation_id and robot_code:
        payload = {
            "robotCode": robot_code,
            "openConversationId": open_conversation_id,
            "msgKey": "sampleFile",
            "msgParam": json.dumps({
                "mediaId": media_id,
                "fileName": file_path.name,
                "fileType": "xlsx",
            }, ensure_ascii=False),
        }
        try:
            body = dingtalk_request_json(
                "https://api.dingtalk.com/v1.0/robot/groupMessages/send",
                payload,
                headers={
                    "Content-Type": "application/json",
                    "x-acs-dingtalk-access-token": access_token,
                },
                timeout=20,
                retries=2,
            )
        except RuntimeError as exc:
            return {"success": False, "message": str(exc), "mediaId": media_id, "target": "openConversationId"}
        success = dingtalk_response_success(body)
        return {"success": success, "body": body, "mediaId": media_id, "target": "openConversationId"}

    chat_id = str(settings.get("dingtalkChatId") or "").strip()
    if not chat_id:
        return {"success": False, "skipped": True, "message": "未配置 openConversationId/robotCode 或 chatId，无法直接发送文件"}
    payload = {
        "chatid": chat_id,
        "msg": {
            "msgtype": "file",
            "file": {"media_id": media_id},
        },
    }
    url = f"https://oapi.dingtalk.com/chat/send?access_token={urllib.parse.quote(access_token)}"
    try:
        body = dingtalk_request_json(
            url,
            payload,
            headers={"Content-Type": "application/json"},
            timeout=20,
            retries=2,
        )
    except RuntimeError as exc:
        return {"success": False, "message": str(exc), "mediaId": media_id}
    return {"success": dingtalk_response_success(body), "body": body, "mediaId": media_id, "target": "chatId"}


def summarize_dingtalk_delivery_failure(result: dict[str, Any]) -> str:
    if not isinstance(result, dict):
        return "未知错误"
    message = result.get("message")
    if not message and isinstance(result.get("body"), dict):
        message = dingtalk_error_message(result["body"], "")
    if not message:
        message = result.get("body") or "未知错误"
    return str(message).replace("\n", " ")[:220]


def build_excel_fallback_markdown(
    excel_path: Path,
    dataset: dict[str, Any],
    direct_result: dict[str, Any],
    base_url: str | None = None,
) -> str:
    excel_link = export_url(excel_path, base_url)
    if is_public_http_url(excel_link):
        file_line = f"- 文件：[{excel_path.name}]({excel_link})"
    else:
        file_line = f"- 文件：{excel_path.name}"
    reason = summarize_dingtalk_delivery_failure(direct_result)
    return "\n".join([
        "### Excel 推荐表下载",
        file_line,
        f"- 范围：{dataset['label']}（{dataset['start']:%Y-%m-%d %H:%M} 至 {dataset['end']:%Y-%m-%d %H:%M}）",
        "- 说明：文件直发未完成，已自动补发下载入口。",
        f"- 直发状态：{reason}",
    ])


def deliver_dingtalk_excel(
    excel_path: Path,
    dataset: dict[str, Any],
    base_url: str | None = None,
) -> dict[str, Any]:
    file_result = send_dingtalk_file(excel_path)
    excel_url = export_url(excel_path, base_url)
    if file_result.get("success"):
        return {
            "success": True,
            "excelDelivery": "dingtalkFile",
            "excelUrl": excel_url,
            "excelFile": file_result,
        }

    if not is_public_http_url(excel_url):
        return {
            "success": False,
            "excelDelivery": "failed",
            "excelUrl": excel_url,
            "excelFile": file_result,
            "message": f"Excel 文件未发送：{summarize_dingtalk_delivery_failure(file_result)}",
        }

    fallback_markdown = build_excel_fallback_markdown(excel_path, dataset, file_result, base_url)
    fallback_result = send_dingtalk_markdown("招聘助手 Excel 推荐表下载", fallback_markdown)
    success = bool(fallback_result.get("success"))
    return {
        "success": success,
        "excelDelivery": "downloadLinkFallback" if success else "failed",
        "excelUrl": excel_url,
        "excelFile": file_result,
        "excelFallback": fallback_result,
        "message": (
            "Excel 文件直发失败，已补发可下载链接到钉钉群"
            if success
            else f"Excel 文件直发失败，下载链接补发也失败：{fallback_result.get('message') or summarize_dingtalk_delivery_failure(file_result)}"
        ),
    }


def parse_dingtalk_message(payload: dict[str, Any]) -> dict[str, Any]:
    """兼容钉钉普通 HTTP 回调与 Stream 包装格式。"""
    if isinstance(payload.get("data"), str):
        try:
            inner = json.loads(payload["data"])
            if isinstance(inner, dict):
                payload = inner
        except json.JSONDecodeError:
            pass

    text = ""
    text_obj = payload.get("text")
    if isinstance(text_obj, dict):
        text = str(text_obj.get("content") or "")
    elif isinstance(text_obj, str):
        text = text_obj
    if not text:
        msg_param = payload.get("msgParam")
        if isinstance(msg_param, str):
            try:
                msg_payload = json.loads(msg_param)
                if isinstance(msg_payload, dict):
                    text = str(msg_payload.get("content") or msg_payload.get("text") or "")
            except json.JSONDecodeError:
                text = msg_param
        if not text:
            text = str(payload.get("content") or payload.get("message") or payload.get("query") or "")

    at_users = payload.get("atUsers") or []
    if isinstance(at_users, list):
        for user in at_users:
            if isinstance(user, dict):
                text = text.replace(str(user.get("dingtalkId") or ""), "")
                text = text.replace(str(user.get("staffId") or ""), "")

    return {
        "question": text.strip(),
        "sessionWebhook": payload.get("sessionWebhook") or "",
        "senderNick": payload.get("senderNick") or payload.get("senderStaffId") or payload.get("senderId") or "",
        "conversationTitle": payload.get("conversationTitle") or "",
        "openConversationId": first_nested_value(payload, "openConversationId", "conversationId", "open_conversation_id"),
        "chatId": first_nested_value(payload, "chatId", "chatid", "chat_id"),
        "robotCode": first_nested_value(payload, "robotCode", "robot_code"),
        "msgtype": payload.get("msgtype") or "text",
        "raw": payload,
    }


def first_nested_value(value: Any, *keys: str) -> str:
    key_set = set(keys)
    if isinstance(value, dict):
        for key in keys:
            found = value.get(key)
            if found:
                return str(found)
        for key, child in value.items():
            if key in key_set and child:
                return str(child)
            found = first_nested_value(child, *keys)
            if found:
                return found
    elif isinstance(value, list):
        for child in value:
            found = first_nested_value(child, *keys)
            if found:
                return found
    return ""


def persist_dingtalk_conversation_target(message: dict[str, Any]) -> dict[str, Any]:
    updates: dict[str, Any] = {}
    if message.get("openConversationId"):
        updates["dingtalkOpenConversationId"] = message["openConversationId"]
    if message.get("chatId"):
        updates["dingtalkChatId"] = message["chatId"]
    if message.get("robotCode"):
        updates["dingtalkRobotCode"] = message["robotCode"]
    if updates:
        save_settings(updates)
    return updates


def dingtalk_callback_ack() -> dict[str, Any]:
    """钉钉 Stream HTTP 推送要求成功响应结构；普通回调也可接受。"""
    return {
        "code": 200,
        "headers": {
            "contentType": "application/json",
            "messageId": f"recruitment_{int(time.time() * 1000)}",
        },
        "message": "OK",
        "data": json.dumps({"response": None}, ensure_ascii=False),
    }


def dingtalk_markdown_reply(title: str, text: str) -> dict[str, Any]:
    return {
        "msgtype": "markdown",
        "markdown": {
            "title": title,
            "text": text,
        },
    }


def dingtalk_stream_ack_with_answer(answer: str) -> dict[str, Any]:
    ack = dingtalk_callback_ack()
    ack["data"] = json.dumps(dingtalk_markdown_reply("招聘助手问答", answer), ensure_ascii=False)
    return ack


def handle_dingtalk_conversation(payload: dict[str, Any]) -> dict[str, Any]:
    message = parse_dingtalk_message(payload)
    saved_target = persist_dingtalk_conversation_target(message)
    question = message["question"]
    if not question:
        answer = "### 招聘助手\n\n我没有收到有效问题。你可以问：昨天推荐了谁？React候选人有哪些？匹配度最高的是谁？"
        agent_meta = {"mode": "empty"}
    else:
        answer, agent_meta = answer_question_with_agent(question)

    save_agent_conversation(
        question=question or "(empty)",
        answer=answer,
        channel="dingtalk",
        sender=message.get("senderNick", ""),
        raw=message.get("raw", {}),
    )

    session_webhook = str(message.get("sessionWebhook") or "").strip()
    if session_webhook and answer:
        try:
            push_result = send_dingtalk_markdown_to_webhook(session_webhook, "招聘助手问答", answer)
        except Exception as exc:
            push_result = {"success": False, "message": f"钉钉 sessionWebhook 回复失败：{exc}"}
    else:
        push_result = {"success": True, "skipped": True, "message": "使用钉钉回调直接回复"}

    return {
        "success": True,
        "question": question,
        "answer": answer,
        "agent": agent_meta,
        "savedTarget": saved_target,
        "reply": push_result,
        "directReply": dingtalk_markdown_reply("招聘助手问答", answer),
        "ack": dingtalk_stream_ack_with_answer(answer),
    }


def build_summary(
    scope: str = "yesterday",
    start: str | None = None,
    end: str | None = None,
    account: str | None = None,
) -> str:
    dataset = get_range_dataset(scope, start, end, account)
    candidates = dataset["candidates"]
    recommendations = dataset["recommendations"]
    received_candidates = dataset["resumeCandidates"]
    recommended_candidates = dataset["recommendedCandidates"]
    title_date = dataset["label"]
    account_names = sorted({
        str(item.get("account_name") or "未识别") for item in [*candidates, *recommendations, *recommended_candidates]
    })

    def account_count(rows: list[dict[str, Any]], account: str) -> int:
        return sum(1 for item in rows if str(item.get("account_name") or "未识别") == account)

    def account_lines(rows: list[dict[str, Any]], label: str, empty: str = "暂无") -> list[str]:
        if not rows:
            return [empty]
        return [
            f"{account}丨{label}：{account_count(rows, account)}"
            for account in account_names
            if account_count(rows, account) > 0
        ] or [empty]

    def metric_line(rows: list[dict[str, Any]], label: str) -> str:
        return " ".join([f"{label}：全部 {len(rows)}", *account_lines(rows, label)])

    source_counts: dict[str, int] = {}
    for item in candidates:
        source = item.get("source") or "未知来源"
        source_counts[source] = source_counts.get(source, 0) + 1
    source_text = "，".join(f"{k} {v}份" for k, v in source_counts.items()) or "暂无"
    return "\n".join(
        [
            "**定时推送**",
            metric_line(candidates, "新增候选人"),
            metric_line(received_candidates, "收到简历数量"),
            metric_line(recommended_candidates, "推荐候选人"),
            f"数据来源：{source_text}",
        ]
    )


def raw_payload(row: dict[str, Any]) -> dict[str, Any]:
    try:
        return json.loads(str(row.get("raw_json") or "{}"))
    except json.JSONDecodeError:
        return {}


def raw_value(row: dict[str, Any], *keys: str) -> str:
    raw = raw_payload(row)
    for key in keys:
        value = row.get(key)
        if value:
            return str(value)
        value = raw.get(key)
        if value:
            return str(value)
    return ""


def has_attachment_resume(row: dict[str, Any]) -> bool:
    raw = raw_payload(row)
    return bool(
        raw.get("hasAttachmentResume")
        or raw.get("resumeAttachmentType") == "attachment"
        or raw.get("resumeEvidence") == "attachmentAccepted"
    )


def account_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        account = str(row.get("account_name") or "未识别")
        counts[account] = counts.get(account, 0) + 1
    return counts


def attachment_resume_count(rows: list[dict[str, Any]]) -> int:
    return sum(1 for row in rows if has_attachment_resume(row))


def attachment_resume_label(row: dict[str, Any]) -> str:
    return "有附件简历" if has_attachment_resume(row) else "无附件简历"


def load_export_manifest() -> dict[str, Any]:
    if not EXPORT_MANIFEST.exists():
        return {}
    try:
        payload = json.loads(EXPORT_MANIFEST.read_text("utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def save_export_manifest(manifest: dict[str, Any]) -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    entries = list(manifest.items())[-500:]
    EXPORT_MANIFEST.write_text(json.dumps(dict(entries), ensure_ascii=False, indent=2), "utf-8")


def register_export(
    path: Path,
    scope: str,
    dataset: dict[str, Any],
    start: str | None = None,
    end: str | None = None,
) -> None:
    manifest = load_export_manifest()
    manifest[path.name] = {
        "filename": path.name,
        "scope": scope,
        "start": start or dataset["start"].isoformat(timespec="seconds"),
        "end": end or dataset["end"].isoformat(timespec="seconds"),
        "label": dataset["label"],
        "account": dataset.get("account") or "",
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "candidateCount": len(dataset.get("candidates") or []),
        "resumeCount": len(dataset.get("resumeCandidates") or []),
        "recommendedCount": len(dataset.get("recommendedCandidates") or []),
    }
    save_export_manifest(manifest)


def sanitize_export_filename(filename: str) -> str:
    name = Path(filename).name
    if not EXPORT_FILENAME_PATTERN.match(name):
        raise ValueError("非法导出文件名")
    return name


def create_recommendation_excel(
    scope: str = "configured",
    start: str | None = None,
    end: str | None = None,
    account: str | None = None,
    output_filename: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception as exc:
        raise RuntimeError(f"Excel 模块不可用，请安装 openpyxl：{exc}") from exc

    dataset = get_range_dataset(scope, start, end, account)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = sanitize_export_filename(output_filename) if output_filename else f"候选人推荐表_{timestamp}.xlsx"
    output_path = EXPORT_DIR / filename

    if RECOMMENDATION_TEMPLATE.exists():
        workbook = load_workbook(RECOMMENDATION_TEMPLATE)
    else:
        workbook = Workbook()
        workbook.active.title = "定时推送数据底表"
        workbook.create_sheet("定时推送候选人详情表模板")

    summary_sheet = workbook["定时推送数据底表"] if "定时推送数据底表" in workbook.sheetnames else workbook.worksheets[0]
    detail_sheet = workbook["定时推送候选人详情表模板"] if "定时推送候选人详情表模板" in workbook.sheetnames else workbook.worksheets[-1]

    for sheet in (summary_sheet, detail_sheet):
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)

    header_fill = PatternFill("solid", fgColor="FFEE00")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="111111")

    summary_headers = ["类目（推荐表中不显示）", "显示名称", "账号", "数量"]
    summary_sheet.append(summary_headers)
    metrics = [
        ("查看候选人计数", "查看候选人", dataset["candidates"]),
        ("收到简历计数", "收到简历", dataset["resumeCandidates"]),
        ("推荐候选人计数", "推荐候选人", dataset["recommendedCandidates"]),
        ("收到附件简历计数", "收到附件简历", [row for row in dataset["resumeCandidates"] if attachment_resume_count([row])]),
        ("收到推荐候选人附件简历计数", "收到推荐候选人附件简历", [row for row in dataset["recommendedCandidates"] if attachment_resume_count([row])]),
    ]
    for metric_key, display_name, rows in metrics:
        summary_sheet.append([metric_key, display_name, "全部", len(rows)])
        for account, count in sorted(account_counts(rows).items()):
            summary_sheet.append([metric_key, display_name, account, count])
    summary_sheet.append(["推送时间范围", dataset["label"], f"{dataset['start']:%Y-%m-%d %H:%M}", f"{dataset['end']:%Y-%m-%d %H:%M}"])

    detail_headers = ["候选人姓名", "学历", "学校", "经验", "是否有附件简历", "来源", "账号", "投递岗位", "匹配度", "匹配度依据"]
    detail_sheet.append(detail_headers)
    for row in sorted(dataset["recommendedCandidates"], key=lambda item: int(item.get("score") or 0), reverse=True):
        raw = raw_payload(row)
        basis = "；".join(
            item for item in [
                raw_value(row, "recommendation"),
                raw_value(row, "next_step", "nextStep"),
                "优势：" + "、".join(raw.get("strengths") or []) if isinstance(raw.get("strengths"), list) and raw.get("strengths") else "",
                "风险：" + "、".join(raw.get("risks") or []) if isinstance(raw.get("risks"), list) and raw.get("risks") else "",
            ]
            if item
        )
        detail_sheet.append([
            raw_value(row, "name") or "未识别",
            raw_value(row, "education"),
            raw_value(row, "school", "schoolName"),
            raw_value(row, "experience"),
            attachment_resume_label(row),
            raw_value(row, "source") or "BOSS直聘",
            raw_value(row, "account_name", "accountName") or "未识别",
            raw_value(row, "role"),
            f"{int(row.get('score') or 0)}%",
            basis[:1000],
        ])

    for sheet in (summary_sheet, detail_sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        sheet.freeze_panes = "A2"

    for column, width in {"A": 24, "B": 18, "C": 20, "D": 12}.items():
        summary_sheet.column_dimensions[column].width = width
    detail_widths = {"A": 16, "B": 10, "C": 18, "D": 12, "E": 14, "F": 12, "G": 16, "H": 28, "I": 10, "J": 70}
    for column, width in detail_widths.items():
        detail_sheet.column_dimensions[column].width = width
    detail_sheet.auto_filter.ref = f"A1:J{max(1, detail_sheet.max_row)}"
    summary_sheet.auto_filter.ref = f"A1:D{max(1, summary_sheet.max_row)}"

    workbook.save(output_path)
    register_export(output_path, scope, dataset, start, end)
    return output_path, dataset


def answer_question_rules(question: str, account: str | None = None) -> str:
    question = question.strip()
    if not question:
        return "请提出一个和招聘历史数据有关的问题。"
    today = date_str()
    yesterday = date_str(-1)
    scope_label = "全部历史"
    scope_date = None
    if "今天" in question or "今日" in question:
        scope_date = today
        scope_label = "今日"
    elif "昨天" in question or "昨日" in question:
        scope_date = yesterday
        scope_label = "昨日"
    filters = {"account": str(account or "").strip()} if str(account or "").strip() else {}
    candidates = list_rows("candidates", limit=1000, date_field="received_date", date_value=scope_date, filters=filters)
    recommendations = list_recommendation_details(limit=1000, date_value=scope_date, filters=filters)
    if account:
        scope_label = f"{scope_label}｜账号：{account}"
    if any(word in question for word in ["汇总", "统计", "多少", "数量"]):
        sources: dict[str, int] = {}
        accounts: dict[str, int] = {}
        for row in candidates:
            sources[row.get("source") or "未知来源"] = sources.get(row.get("source") or "未知来源", 0) + 1
            accounts[row.get("account_name") or "未识别"] = accounts.get(row.get("account_name") or "未识别", 0) + 1
        source_text = "，".join(f"{k} {v}份" for k, v in sources.items()) or "暂无"
        account_text = "，".join(f"{k} {v}份" for k, v in accounts.items()) or "暂无"
        return f"### 招聘助手答复\n\n- 查询范围：{scope_label}\n- 候选人数量：{len(candidates)}\n- 推荐候选人：{len(recommendations)}\n- 数据来源：{source_text}\n- 账号分布：{account_text}"
    cleaned = question
    highest_query = any(word in question for word in ["最高", "最好", "最合适", "排名"])
    for word in ["今天", "今日", "昨天", "昨日", "候选人", "推荐", "简历", "哪些", "哪个", "有没有", "最高", "最好", "最合适", "匹配度", "统计", "汇总", "来源", "账号", "的是谁", "是谁", "的", "谁"]:
        cleaned = cleaned.replace(word, " ")
    keywords = [word for word in cleaned.replace("？", " ").replace("?", " ").split() if len(word) >= 2]
    rows = recommendations or candidates
    if keywords and not highest_query:
        rows = [
            row for row in rows
            if any(keyword.lower() in json.dumps(dict(row), ensure_ascii=False).lower() for keyword in keywords)
        ]
    rows = sorted(rows, key=lambda row: row.get("score") or 0, reverse=True)
    if not rows:
        return f"### 招聘助手答复\n\n没有找到和“{question}”匹配的历史候选人。"
    lines = [
        f"{idx + 1}. {row.get('name','未识别')}｜{row.get('role','待确认')}｜{row.get('score',0)}%｜{row.get('recommendation','待评估')}｜{row.get('source','未知来源')}｜{row.get('account_name','未识别')}"
        for idx, row in enumerate(rows[:10])
    ]
    role_requirements = []
    seen_roles = set()
    for row in rows[:5]:
        role = row.get("role") or ""
        normalized = normalize_role(role)
        if not normalized or normalized in seen_roles:
            continue
        seen_roles.add(normalized)
        requirement = get_job_requirement(role, account or "")
        if requirement:
            role_requirements.append(f"- {role}：{str(requirement.get('requirement') or '')[:180]}")
    requirement_text = "\n\n岗位要求依据：\n" + "\n".join(role_requirements) if role_requirements else ""
    return f"### 招聘助手答复\n\n查询范围：{scope_label}\n\n" + "\n".join(lines) + requirement_text


def recover_missing_export(filename: str) -> Path | None:
    try:
        safe_name = sanitize_export_filename(filename)
    except ValueError:
        return None
    manifest = load_export_manifest()
    meta = manifest.get(safe_name) if isinstance(manifest.get(safe_name), dict) else {}
    scope = str(meta.get("scope") or "configured")
    start = str(meta.get("start") or "").strip() or None
    end = str(meta.get("end") or "").strip() or None
    account = str(meta.get("account") or "").strip() or None
    try:
        path, dataset = create_recommendation_excel(scope, start, end, account, output_filename=safe_name)
    except Exception as exc:
        print(f"[exports] recover failed: {safe_name}: {exc}")
        return None
    if path.exists() and path.is_file():
        print(f"[exports] recovered missing Excel: {safe_name} scope={scope} label={dataset['label']}")
        return path
    return None


class Handler(SimpleHTTPRequestHandler):
    server_version = "HRassistant"
    sys_version = ""

    def end_headers(self) -> None:
        if not getattr(self, "_common_headers_written", False):
            write_common_headers(self)
        super().end_headers()

    def translate_path(self, path: str) -> str:
        parsed = urllib.parse.urlparse(path)
        if parsed.path == "/":
            return str(STATIC_DIR / "index.html")
        return str(STATIC_DIR / parsed.path.lstrip("/"))

    def serve_extension_package(self, query: dict[str, list[str]]) -> None:
        backend_url = configured_backend_url(self, query.get("backendUrl", [""])[0])
        data, meta = build_configured_extension_package(backend_url)
        binary_response(
            self,
            data,
            "application/zip",
            str(meta["filename"]),
        )

    def serve_export(self, path_value: str, head_only: bool = False) -> bool:
        filename = Path(urllib.parse.unquote(path_value.split("/exports/", 1)[1])).name
        path = EXPORT_DIR / filename
        if not path.exists() or not path.is_file():
            recovered = recover_missing_export(filename)
            if not recovered:
                json_response(self, {"success": False, "message": "文件不存在且无法自动恢复"}, 404)
                return True
            path = recovered
        binary_response(
            self,
            path.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            path.name,
            head_only=head_only,
        )
        return True

    def do_OPTIONS(self) -> None:
        if not guard_request(self):
            return
        json_response(self, {"success": True})

    def do_HEAD(self) -> None:
        if not guard_request(self):
            return
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path.startswith("/exports/"):
            self.serve_export(parsed.path, head_only=True)
            return
        super().do_HEAD()

    def do_GET(self) -> None:
        if not guard_request(self):
            return
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)
        try:
            if parsed.path == "/api/health":
                settings = get_settings()
                json_response(self, {
                    "success": True,
                    "time": now_iso(),
                    "callback": "/api/dingtalk/callback",
                    "dingtalkTargetConfigured": bool(settings.get("dingtalkOpenConversationId") or settings.get("dingtalkChatId")),
                    "security": {
                        "allowedHosts": ALLOWED_HOSTS,
                        "allowedOrigins": ALLOWED_ORIGINS,
                        "clientIpAllowlistEnabled": bool(ALLOWED_CLIENT_IPS),
                        "rateLimitPerMinute": RATE_LIMIT_PER_MINUTE,
                    },
                })
            elif parsed.path == "/api/security/allowlist":
                json_response(self, get_security_allowlist_config(self))
            elif parsed.path == "/api/extension/config":
                backend_url = configured_backend_url(self, query.get("backendUrl", [""])[0])
                _, meta = build_configured_extension_package(backend_url)
                json_response(self, {"success": True, **meta})
            elif parsed.path == "/api/extension/package":
                self.serve_extension_package(query)
            elif parsed.path == "/api/stats":
                json_response(self, get_stats({
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                }))
            elif parsed.path == "/api/accounts":
                json_response(self, {"items": get_accounts()})
            elif parsed.path == "/api/settings":
                settings = get_settings()
                if not usable_base_url(settings.get("publicBaseUrl")):
                    settings["publicBaseUrl"] = DEFAULT_PUBLIC_BASE_URL
                if not usable_base_url(settings.get("adminBaseUrl")):
                    settings["adminBaseUrl"] = DEFAULT_PUBLIC_BASE_URL
                llm_config = get_llm_config(mask_key=True)
                settings.update(llm_config)
                if settings.get("dingtalkAppKey"):
                    settings["dingtalkAppKeyConfigured"] = True
                    settings["dingtalkAppKey"] = "********"
                if settings.get("dingtalkAppSecret"):
                    settings["dingtalkAppSecretConfigured"] = True
                    settings["dingtalkAppSecret"] = "********"
                json_response(self, settings)
            elif parsed.path == "/api/llm/config":
                json_response(self, get_llm_config(mask_key=True))
            elif parsed.path == "/api/behavior-policy":
                json_response(self, get_behavior_policy(query.get("account", [""])[0]))
            elif parsed.path == "/api/candidates":
                json_response(self, {"items": list_rows("candidates", int(query.get("limit", ["200"])[0]), filters={
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                })})
            elif parsed.path == "/api/recommendations":
                json_response(self, {"items": list_recommendation_details(int(query.get("limit", ["200"])[0]), filters={
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                })})
            elif parsed.path == "/api/reports":
                json_response(self, {"items": list_rows("reports", int(query.get("limit", ["100"])[0]), filters={
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                })})
            elif parsed.path == "/api/job-requirements":
                role = query.get("role", [""])[0]
                account = query.get("account", [""])[0]
                if role:
                    json_response(self, {"item": get_job_requirement(role, account)})
                else:
                    json_response(self, {"items": list_rows("job_requirements", int(query.get("limit", ["200"])[0]), filters={
                        "q": query.get("q", [""])[0],
                        "source": query.get("source", [""])[0],
                        "account": account,
                        "accountExact": query.get("accountExact", [""])[0],
                    })})
            elif parsed.path == "/api/agent/conversations":
                json_response(self, {"items": list_agent_conversations(int(query.get("limit", ["100"])[0]))})
            elif parsed.path == "/api/summary":
                scope = query.get("scope", ["all"])[0]
                json_response(self, {"markdown": build_summary(
                    scope,
                    query.get("start", [""])[0] or None,
                    query.get("end", [""])[0] or None,
                    query.get("account", [""])[0] or None,
                )})
            elif parsed.path == "/api/summary/excel":
                scope = query.get("scope", ["configured"])[0]
                output_path, dataset = create_recommendation_excel(
                    scope,
                    query.get("start", [""])[0] or None,
                    query.get("end", [""])[0] or None,
                    query.get("account", [""])[0] or None,
                )
                binary_response(
                    self,
                    output_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    output_path.name,
                )
            elif parsed.path.startswith("/exports/"):
                self.serve_export(parsed.path)
            elif parsed.path == "/api/export/candidates.csv":
                rows = list_rows("candidates", 10000, filters={
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                })
                text_response(self, rows_to_csv(rows, [
                    ("received_date", "日期"),
                    ("name", "姓名"),
                    ("role", "岗位"),
                    ("education", "学历"),
                    ("experience", "经验"),
                    ("expected_salary", "薪资"),
                    ("score", "匹配度"),
                    ("recommendation", "推荐意见"),
                    ("source", "数据来源"),
                    ("account_name", "账号信息"),
                    ("created_at", "创建时间"),
                ]), content_type="text/csv; charset=utf-8", filename="candidates.csv")
            elif parsed.path == "/api/export/recommendations.csv":
                rows = list_recommendation_details(10000, filters={
                    "q": query.get("q", [""])[0],
                    "source": query.get("source", [""])[0],
                    "account": query.get("account", [""])[0],
                    "accountExact": query.get("accountExact", [""])[0],
                })
                text_response(self, rows_to_csv(rows, [
                    ("created_at", "推荐时间"),
                    ("name", "姓名"),
                    ("role", "岗位"),
                    ("score", "匹配度"),
                    ("recommendation", "推荐意见"),
                    ("next_step", "下一步"),
                    ("source", "数据来源"),
                    ("account_name", "账号信息"),
                ]), content_type="text/csv; charset=utf-8", filename="recommendations.csv")
            else:
                super().do_GET()
        except ValueError as exc:
            json_response(self, {"success": False, "message": str(exc)}, 413)
        except Exception as exc:
            json_response(self, {"success": False, "message": str(exc)}, 500)

    def do_POST(self) -> None:
        if not guard_request(self):
            return
        parsed = urllib.parse.urlparse(self.path)
        try:
            payload = read_json(self)
            if parsed.path == "/api/settings":
                json_response(self, save_settings(payload))
            elif parsed.path == "/api/accounts":
                action = str(payload.get("action") or "save").strip().lower()
                if action == "delete":
                    json_response(self, delete_managed_account(payload))
                else:
                    json_response(self, save_managed_account(payload))
            elif parsed.path == "/api/security/allowlist":
                json_response(self, save_security_allowlist_config(payload, self))
            elif parsed.path == "/api/llm/config":
                json_response(self, save_llm_config(payload))
            elif parsed.path == "/api/llm/config/reset":
                json_response(self, reset_llm_config())
            elif parsed.path == "/api/behavior-policy":
                json_response(self, save_behavior_policy(payload))
            elif parsed.path == "/api/page-intelligence/extract":
                json_response(self, extract_page_intelligence(payload))
            elif parsed.path == "/api/candidates":
                json_response(self, upsert_candidate(payload))
            elif parsed.path == "/api/recommendations":
                json_response(self, save_recommendation(payload.get("candidate", payload), payload.get("report", "")))
            elif parsed.path == "/api/job-requirements":
                json_response(self, upsert_job_requirement(payload))
            elif parsed.path == "/api/job-requirements/pending":
                json_response(self, upsert_pending_job_requirement(payload))
            elif parsed.path == "/api/job-requirements/match-candidates":
                account_exact_value = payload.get("accountExact", payload.get("account_exact", "true"))
                json_response(self, match_candidates_with_job_requirements(
                    str(payload.get("role", "")),
                    str(payload.get("account", "")),
                    str(account_exact_value).strip().lower() in {"1", "true", "yes", "on"},
                ))
            elif parsed.path == "/api/agent/ask":
                answer, agent_meta = answer_question_with_agent(
                    str(payload.get("question", "")),
                    str(payload.get("account", "") or "").strip() or None,
                    bool(payload.get("forceLlm")),
                )
                save_agent_conversation(
                    question=str(payload.get("question", "")),
                    answer=answer,
                    channel="web",
                    sender=str(payload.get("sender", "Web 管理后台")),
                    raw=payload,
                )
                if payload.get("replyToDingTalk"):
                    send_dingtalk_markdown("招聘助手问答", answer)
                json_response(self, {"success": True, "answer": answer, "agent": agent_meta})
            elif parsed.path == "/api/dingtalk/test":
                json_response(self, send_dingtalk_markdown("招聘助手钉钉连接测试", "### 招聘助手钉钉连接测试\n\n连接成功。"))
            elif parsed.path == "/api/summary/push":
                query = urllib.parse.parse_qs(parsed.query)
                scope = query.get("scope", ["configured"])[0]
                base_url = request_base_url(self)
                settings = get_settings()
                if usable_base_url(settings.get("publicBaseUrl")) != str(settings.get("publicBaseUrl") or "").strip().rstrip("/"):
                    save_settings({"publicBaseUrl": ""})
                if not usable_base_url(settings.get("publicBaseUrl")):
                    save_settings({"publicBaseUrl": DEFAULT_PUBLIC_BASE_URL})
                if not usable_base_url(settings.get("adminBaseUrl")):
                    save_settings({"adminBaseUrl": DEFAULT_PUBLIC_BASE_URL})
                output_path, dataset = create_recommendation_excel(
                    scope,
                    query.get("start", [""])[0] or None,
                    query.get("end", [""])[0] or None,
                    query.get("account", [""])[0] or None,
                )
                markdown = build_push_markdown(scope, dataset, output_path, base_url)
                summary_result = send_dingtalk_markdown("招聘助手候选人简历汇总", markdown)
                delivery_result = deliver_dingtalk_excel(output_path, dataset, base_url)
                excel_url = export_url(output_path, base_url)
                result = dict(summary_result)
                summary_ok = bool(summary_result.get("success"))
                delivery_ok = bool(delivery_result.get("success"))
                result["success"] = summary_ok and delivery_ok
                result["excelDelivery"] = delivery_result.get("excelDelivery")
                if result["success"] and delivery_result.get("excelDelivery") == "dingtalkFile":
                    result["message"] = "汇总正文和 Excel 文件已同步至钉钉。"
                elif result["success"]:
                    result["message"] = delivery_result.get("message") or "汇总正文已发送，Excel 下载入口已同步至钉钉。"
                elif summary_ok:
                    result["message"] = delivery_result.get("message") or "正文已发送，但 Excel 文件未同步至钉钉。"
                elif delivery_ok:
                    result["message"] = f"Excel 已同步至钉钉，但正文发送失败：{summary_result.get('message') or summary_result.get('body') or '未知错误'}"
                else:
                    result["message"] = (
                        f"正文发送失败：{summary_result.get('message') or summary_result.get('body') or '未知错误'}；"
                        f"Excel 同步失败：{delivery_result.get('message') or delivery_result.get('excelFile') or '未知错误'}"
                    )
                result.update({
                    "excel": output_path.name,
                    "excelUrl": excel_url,
                    "excelFile": delivery_result.get("excelFile"),
                    "excelFallback": delivery_result.get("excelFallback"),
                    "range": {
                        "label": dataset["label"],
                        "start": dataset["start"].isoformat(timespec="seconds"),
                        "end": dataset["end"].isoformat(timespec="seconds"),
                    },
                })
                json_response(self, result)
            elif parsed.path == "/api/dingtalk/callback":
                result = handle_dingtalk_conversation(payload)
                if payload.get("data") is not None:
                    json_response(self, result["ack"])
                else:
                    json_response(self, result["directReply"])
            elif parsed.path == "/api/dingtalk/callback-test":
                question = str(payload.get("question") or "")
                answer, agent_meta = answer_question_with_agent(question)
                save_agent_conversation(question=question, answer=answer, channel="test", sender="callback-test", raw=payload)
                json_response(self, {"success": True, "question": question, "answer": answer, "agent": agent_meta})
            else:
                json_response(self, {"success": False, "message": "not found"}, 404)
        except ValueError as exc:
            json_response(self, {"success": False, "message": str(exc)}, 413)
        except Exception as exc:
            json_response(self, {"success": False, "message": str(exc)}, 500)


def main() -> None:
    init_db()
    threading.Thread(target=scheduled_push_loop, daemon=True).start()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"招聘助手 Web 管理后台已启动: http://127.0.0.1:{PORT}")
    print(f"数据文件: {DB_PATH}")
    server.serve_forever()


if __name__ == "__main__":
    main()
